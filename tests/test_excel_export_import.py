"""
Test Excel Export/Import functionality for Balia prices
Tests:
- GET /api/prices/export - Download Excel file with 3 sheets (Modele, Opcje, Ustawienia)
- POST /api/prices/import - Upload and update prices from Excel file
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestExcelExport:
    """Test Excel export endpoint"""
    
    def test_export_returns_excel_file(self):
        """Test that export endpoint returns an Excel file"""
        response = requests.get(f"{BASE_URL}/api/prices/export")
        
        # Status code assertion
        assert response.status_code == 200, f"Expected 200, got {response.status_code}"
        
        # Content type assertion - should be Excel
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheet' in content_type or 'excel' in content_type or 'octet-stream' in content_type, \
            f"Expected Excel content type, got {content_type}"
        
        # Content-Disposition should have filename
        content_disposition = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disposition, f"Expected attachment, got {content_disposition}"
        assert '.xlsx' in content_disposition, f"Expected .xlsx filename, got {content_disposition}"
        
        # File should have content
        assert len(response.content) > 0, "Excel file should not be empty"
        
        print(f"✓ Export returns Excel file: {len(response.content)} bytes")
        print(f"  Content-Type: {content_type}")
        print(f"  Content-Disposition: {content_disposition}")
    
    def test_export_file_has_valid_excel_structure(self):
        """Test that exported file is a valid Excel with expected sheets"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        response = requests.get(f"{BASE_URL}/api/prices/export")
        assert response.status_code == 200
        
        # Load workbook from response content
        buffer = io.BytesIO(response.content)
        wb = load_workbook(buffer, read_only=True)
        
        # Check sheet names
        expected_sheets = ["Modele", "Opcje", "Ustawienia"]
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"
        
        print(f"✓ Excel has all expected sheets: {wb.sheetnames}")
        
        # Check Modele sheet has headers
        ws_models = wb["Modele"]
        headers = [cell.value for cell in next(ws_models.iter_rows(min_row=1, max_row=1))]
        expected_model_headers = ["ID", "Nazwa (RU)", "Nazwa (PL)", "Typ печi", "Zakup EUR", "Marża %", "Cena PLN", "Kolor HEX"]
        for header in expected_model_headers:
            assert header in headers, f"Missing header in Modele: {header}"
        print(f"✓ Modele sheet has correct headers: {headers}")
        
        # Check Opcje sheet has headers
        ws_options = wb["Opcje"]
        option_headers = [cell.value for cell in next(ws_options.iter_rows(min_row=1, max_row=1))]
        expected_option_headers = ["Kategoria ID", "Kategoria (RU)", "Opcja ID", "Opcja (RU)", "Opcja (PL)", "Zakup EUR", "Marża %", "Cena PLN", "Kolor HEX"]
        for header in expected_option_headers:
            assert header in option_headers, f"Missing header in Opcje: {header}"
        print(f"✓ Opcje sheet has correct headers: {option_headers}")
        
        # Check Ustawienia sheet has settings
        ws_settings = wb["Ustawienia"]
        settings_params = []
        for row in ws_settings.iter_rows(min_row=2, values_only=True):
            if row[0]:
                settings_params.append(row[0])
        
        expected_settings = ["Waluta", "Symbol waluty", "Kurs EUR", "Domyślna marża %"]
        for setting in expected_settings:
            assert setting in settings_params, f"Missing setting: {setting}"
        print(f"✓ Ustawienia sheet has correct settings: {settings_params}")
        
        wb.close()


class TestExcelImport:
    """Test Excel import endpoint"""
    
    def test_import_rejects_non_excel_file(self):
        """Test that import rejects non-Excel files"""
        # Create a fake text file
        files = {'file': ('test.txt', b'This is not an Excel file', 'text/plain')}
        response = requests.post(f"{BASE_URL}/api/prices/import", files=files)
        
        # Should reject with 400
        assert response.status_code == 400, f"Expected 400 for non-Excel file, got {response.status_code}"
        print(f"✓ Import correctly rejects non-Excel files: {response.json()}")
    
    def test_import_accepts_exported_file(self):
        """Test that import accepts a file exported from the system"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # First export the current prices
        export_response = requests.get(f"{BASE_URL}/api/prices/export")
        assert export_response.status_code == 200, "Export failed"
        
        # Import the same file back
        files = {'file': ('cennik.xlsx', export_response.content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        import_response = requests.post(f"{BASE_URL}/api/prices/import", files=files)
        
        # Should succeed
        assert import_response.status_code == 200, f"Import failed: {import_response.text}"
        
        data = import_response.json()
        assert 'message' in data, "Response should have message"
        assert 'updated_models' in data, "Response should have updated_models count"
        assert 'updated_options' in data, "Response should have updated_options count"
        assert 'updated_settings' in data, "Response should have updated_settings count"
        
        print(f"✓ Import successful: {data}")
    
    def test_import_updates_prices(self):
        """Test that import actually updates prices in database"""
        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Get current prices
        prices_before = requests.get(f"{BASE_URL}/api/prices").json()
        
        # Export current prices
        export_response = requests.get(f"{BASE_URL}/api/prices/export")
        assert export_response.status_code == 200
        
        # Modify the exported file - change EUR rate in Ustawienia
        buffer = io.BytesIO(export_response.content)
        wb = load_workbook(buffer)
        
        ws_settings = wb["Ustawienia"]
        original_rate = None
        test_rate = 5.55  # Test value
        
        for row in ws_settings.iter_rows(min_row=2):
            if row[0].value == "Kurs EUR":
                original_rate = row[1].value
                row[1].value = test_rate
                break
        
        # Save modified workbook
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()
        
        # Import modified file
        files = {'file': ('cennik_modified.xlsx', output.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        import_response = requests.post(f"{BASE_URL}/api/prices/import", files=files)
        assert import_response.status_code == 200, f"Import failed: {import_response.text}"
        
        # Verify the change was applied
        prices_after = requests.get(f"{BASE_URL}/api/prices").json()
        assert prices_after.get('eurRate') == test_rate, f"EUR rate not updated: expected {test_rate}, got {prices_after.get('eurRate')}"
        
        print(f"✓ Import updates prices: EUR rate changed from {original_rate} to {test_rate}")
        
        # Restore original rate
        if original_rate is not None:
            restore_data = {**prices_after, 'eurRate': original_rate}
            requests.post(f"{BASE_URL}/api/prices", json=restore_data)
            print(f"✓ Restored original EUR rate: {original_rate}")


class TestExportImportRoundTrip:
    """Test full export-import cycle"""
    
    def test_export_import_preserves_data(self):
        """Test that export->import cycle preserves all data"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Get current prices
        prices_before = requests.get(f"{BASE_URL}/api/prices").json()
        
        # Export
        export_response = requests.get(f"{BASE_URL}/api/prices/export")
        assert export_response.status_code == 200
        
        # Import same file
        files = {'file': ('cennik.xlsx', export_response.content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        import_response = requests.post(f"{BASE_URL}/api/prices/import", files=files)
        assert import_response.status_code == 200
        
        # Get prices after
        prices_after = requests.get(f"{BASE_URL}/api/prices").json()
        
        # Compare key fields
        assert prices_before.get('currency') == prices_after.get('currency'), "Currency changed"
        assert prices_before.get('currencySymbol') == prices_after.get('currencySymbol'), "Currency symbol changed"
        assert prices_before.get('eurRate') == prices_after.get('eurRate'), "EUR rate changed"
        
        # Compare model count
        assert len(prices_before.get('models', [])) == len(prices_after.get('models', [])), "Model count changed"
        
        # Compare category count
        assert len(prices_before.get('categories', [])) == len(prices_after.get('categories', [])), "Category count changed"
        
        print(f"✓ Export-Import round trip preserves data")
        print(f"  Models: {len(prices_after.get('models', []))}")
        print(f"  Categories: {len(prices_after.get('categories', []))}")
        print(f"  Currency: {prices_after.get('currency')}")
        print(f"  EUR Rate: {prices_after.get('eurRate')}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
