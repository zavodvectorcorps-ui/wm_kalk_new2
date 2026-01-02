"""
Test suite for POST /api/generate-production-excel endpoint
Tests Excel generation with marked options (X) for Balia orders
"""
import pytest
import requests
import os
import io
from openpyxl import load_workbook

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestProductionExcelGeneration:
    """Tests for production Excel generation endpoint"""
    
    def test_endpoint_exists_and_returns_excel(self):
        """Test that endpoint returns Excel file"""
        payload = {
            "fullName": "Test Customer",
            "fullAddress": "Test Address 123",
            "orderId": "TEST-001"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200, f"Expected 200, got {response.status_code}: {response.text}"
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.headers.get('Content-Type', '')
        assert 'attachment' in response.headers.get('Content-Disposition', '')
        assert '.xlsx' in response.headers.get('Content-Disposition', '')
        print("PASS: Endpoint returns Excel file with correct headers")
    
    def test_customer_data_in_excel(self):
        """Test that customer data (fullName, fullAddress) is written to B2 and B4"""
        payload = {
            "fullName": "Jan Kowalski",
            "fullAddress": "ul. Testowa 15, Warszawa",
            "orderId": "TEST-002"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        
        # Load Excel from response
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check customer data
        assert ws['B2'].value == "Jan Kowalski", f"B2 should be 'Jan Kowalski', got '{ws['B2'].value}'"
        assert ws['B4'].value == "ul. Testowa 15, Warszawa", f"B4 should be address, got '{ws['B4'].value}'"
        print("PASS: Customer data written to B2 (name) and B4 (address)")
    
    def test_heater_type_external_marked(self):
        """Test that external heater type is marked with X in B10"""
        payload = {
            "fullName": "Test",
            "fullAddress": "Test",
            "selectedHeaterType": "external",
            "orderId": "TEST-003"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        assert ws['B10'].value == 'X', f"B10 should be 'X' for external heater, got '{ws['B10'].value}'"
        assert ws['C10'].value is None or ws['C10'].value != 'X', "C10 should NOT be marked for external heater"
        print("PASS: External heater type marked with X in B10")
    
    def test_heater_type_integrated_marked(self):
        """Test that integrated heater type is marked with X in C10"""
        payload = {
            "fullName": "Test",
            "fullAddress": "Test",
            "selectedHeaterType": "integrated",
            "orderId": "TEST-004"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        assert ws['C10'].value == 'X', f"C10 should be 'X' for integrated heater, got '{ws['C10'].value}'"
        assert ws['B10'].value is None or ws['B10'].value != 'X', "B10 should NOT be marked for integrated heater"
        print("PASS: Integrated heater type marked with X in C10")
    
    def test_fiberglass_color_marked(self):
        """Test that fiberglass colors are marked correctly in D10-R10"""
        # Test fg_white -> D10
        payload = {
            "fullName": "Test",
            "fullAddress": "Test",
            "selections": {
                "fiberglass_color": "fg_white"
            },
            "orderId": "TEST-005"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        assert ws['D10'].value == 'X', f"D10 should be 'X' for fg_white, got '{ws['D10'].value}'"
        print("PASS: Fiberglass color fg_white marked with X in D10")
    
    def test_acrylic_color_marked(self):
        """Test that acrylic colors are marked correctly in V10-AB10"""
        # Test ac_white -> V10
        payload = {
            "fullName": "Test",
            "fullAddress": "Test",
            "selections": {
                "acrylic_color": "ac_white"
            },
            "orderId": "TEST-006"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        assert ws['V10'].value == 'X', f"V10 should be 'X' for ac_white, got '{ws['V10'].value}'"
        print("PASS: Acrylic color ac_white marked with X in V10")
    
    def test_model_marked(self):
        """Test that model is marked correctly in Y16-AD16"""
        # Test round_200 -> Y16
        payload = {
            "fullName": "Test",
            "fullAddress": "Test",
            "modelId": "round_200",
            "orderId": "TEST-007"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        assert ws['Y16'].value == 'X', f"Y16 should be 'X' for round_200, got '{ws['Y16'].value}'"
        print("PASS: Model round_200 marked with X in Y16")
    
    def test_accessories_marked(self):
        """Test that accessories are marked correctly in B16-V16"""
        # Test multiple accessories via selections dict
        payload = {
            "fullName": "Test",
            "fullAddress": "Test",
            "selections": {
                "accessories": {
                    "hydromassage_basic": True,
                    "air_bubble_yes": True,
                    "head_pillow": True
                }
            },
            "orderId": "TEST-008"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        assert ws['B16'].value == 'X', f"B16 should be 'X' for hydromassage_basic, got '{ws['B16'].value}'"
        assert ws['C16'].value == 'X', f"C16 should be 'X' for air_bubble_yes, got '{ws['C16'].value}'"
        assert ws['K16'].value == 'X', f"K16 should be 'X' for head_pillow, got '{ws['K16'].value}'"
        print("PASS: Accessories marked correctly in row 16")
    
    def test_notes_written(self):
        """Test that notes are written to B18"""
        payload = {
            "fullName": "Test",
            "fullAddress": "Test",
            "notes": "Special instructions for production",
            "orderId": "TEST-009"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        assert ws['B18'].value == "Special instructions for production", f"B18 should contain notes, got '{ws['B18'].value}'"
        print("PASS: Notes written to B18")
    
    def test_full_order_with_all_options(self):
        """Test complete order with heater, model, colors, and accessories"""
        payload = {
            "fullName": "Piotr Nowak",
            "fullAddress": "ul. Główna 10, Kraków",
            "selectedHeaterType": "external",
            "modelId": "round_225",
            "selections": {
                "fiberglass_color": "fg_blue",
                "lid": "glass_fiber_lid",
                "wood_finish": "natural",
                "wood_type": "spruce",
                "accessories": {
                    "insulation_yes": True,
                    "bluetooth_radio": True
                }
            },
            "notes": "Urgent order - deliver by Friday",
            "orderId": "TEST-010"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verify all marks
        assert ws['B2'].value == "Piotr Nowak", "Customer name should be in B2"
        assert ws['B4'].value == "ul. Główna 10, Kraków", "Address should be in B4"
        assert ws['B10'].value == 'X', "External heater should be marked in B10"
        assert ws['Z16'].value == 'X', "round_225 model should be marked in Z16"
        assert ws['F10'].value == 'X', "fg_blue color should be marked in F10"
        assert ws['AD10'].value == 'X', "glass_fiber_lid should be marked in AD10"
        assert ws['AF10'].value == 'X', "natural wood finish should be marked in AF10"
        assert ws['AE16'].value == 'X', "spruce wood type should be marked in AE16"
        assert ws['J16'].value == 'X', "insulation_yes should be marked in J16"
        assert ws['T16'].value == 'X', "bluetooth_radio should be marked in T16"
        assert ws['B18'].value == "Urgent order - deliver by Friday", "Notes should be in B18"
        print("PASS: Full order with all options marked correctly")
    
    def test_filename_contains_customer_name_and_order_id(self):
        """Test that filename contains customer name and order ID"""
        payload = {
            "fullName": "Anna Wiśniewska",
            "fullAddress": "Test",
            "orderId": "WMB-2025-001",
            "id": "backup-id"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        content_disposition = response.headers.get('Content-Disposition', '')
        
        # Filename should contain TechSpec, customer name, and order ID
        assert 'TechSpec' in content_disposition, f"Filename should contain 'TechSpec', got: {content_disposition}"
        assert 'WMB-2025-001' in content_disposition, f"Filename should contain order ID, got: {content_disposition}"
        print(f"PASS: Filename format correct: {content_disposition}")
    
    def test_selected_options_array_backward_compatibility(self):
        """Test that selectedOptions array also marks options (backward compatibility)"""
        payload = {
            "fullName": "Test",
            "fullAddress": "Test",
            "selectedOptions": [
                {"id": "fg_gray"},
                {"id": "sand_filter"},
                {"id": "electric_heater"}
            ],
            "orderId": "TEST-011"
        }
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        
        assert response.status_code == 200
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        assert ws['G10'].value == 'X', f"G10 should be 'X' for fg_gray, got '{ws['G10'].value}'"
        assert ws['L16'].value == 'X', f"L16 should be 'X' for sand_filter, got '{ws['L16'].value}'"
        assert ws['U16'].value == 'X', f"U16 should be 'X' for electric_heater, got '{ws['U16'].value}'"
        print("PASS: selectedOptions array backward compatibility works")


class TestProductionExcelWithRealOrder:
    """Test with real order data from database"""
    
    def test_generate_excel_for_existing_order(self):
        """Test generating Excel for an existing order from the database"""
        # First get an existing order
        orders_response = requests.get(f"{BASE_URL}/api/orders")
        assert orders_response.status_code == 200
        orders = orders_response.json()
        
        if not orders:
            pytest.skip("No existing orders to test with")
        
        # Use first order
        order = orders[0]
        print(f"Testing with order: {order.get('id')} - {order.get('fullName')}")
        
        # Generate Excel for this order
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json={
            **order,
            "orderId": order.get('id')
        })
        
        assert response.status_code == 200, f"Failed to generate Excel: {response.text}"
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.headers.get('Content-Type', '')
        
        # Verify Excel content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Customer data should be present
        if order.get('fullName'):
            assert ws['B2'].value == order['fullName'], f"B2 should be '{order['fullName']}', got '{ws['B2'].value}'"
        if order.get('fullAddress'):
            assert ws['B4'].value == order['fullAddress'], f"B4 should be '{order['fullAddress']}', got '{ws['B4'].value}'"
        
        print(f"PASS: Excel generated for existing order {order.get('id')}")
        print(f"  - Customer: {ws['B2'].value}")
        print(f"  - Address: {ws['B4'].value}")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
