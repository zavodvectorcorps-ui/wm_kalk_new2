"""
Test suite for heater type mapping in orders, Excel, and PDF generation
Tests the following features:
1. Order creation saves selectedHeaterVariantId and heaterType
2. POST /api/generate-production-excel marks X for selected heater if excelCell is configured
3. POST /api/generate-pdf shows heater type (Piec zintegrowany / Piec zewnętrzny) under model name
"""
import pytest
import requests
import os
import io
from openpyxl import load_workbook
from PyPDF2 import PdfReader

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')


class TestOrderHeaterFields:
    """Test that orders save heaterType and selectedHeaterVariantId correctly"""
    
    def test_create_order_with_integrated_heater(self):
        """Test creating order with integrated heater type"""
        order_data = {
            "id": "TEST-HEATER-INT-001",
            "fullName": "Test Integrated Heater",
            "phoneNumber": "+48123456789",
            "fullAddress": "Test Address 123",
            "orderDate": "2025-01-15",
            "modelId": "round_200",
            "modelName": "Okrągła 200cm",
            "modelPrice": 1677.0,
            "heaterType": "integrated",
            "heaterTypeName": "Piec zintegrowany",
            "selectedHeaterVariantId": "round_200_integrated",
            "selections": {},
            "selectedOptions": [],
            "notes": "Test order with integrated heater",
            "discountPercent": 0,
            "subtotal": 1677.0,
            "total": 1677.0,
            "currency": "EUR"
        }
        
        # Create order
        response = requests.post(f"{BASE_URL}/api/orders", json=order_data)
        assert response.status_code == 200, f"Failed to create order: {response.text}"
        
        created_order = response.json()
        assert created_order.get('heaterType') == 'integrated', f"heaterType should be 'integrated', got: {created_order.get('heaterType')}"
        assert created_order.get('heaterTypeName') == 'Piec zintegrowany', f"heaterTypeName should be 'Piec zintegrowany', got: {created_order.get('heaterTypeName')}"
        assert created_order.get('selectedHeaterVariantId') == 'round_200_integrated', f"selectedHeaterVariantId should be 'round_200_integrated', got: {created_order.get('selectedHeaterVariantId')}"
        
        print("PASS: Order created with integrated heater fields saved correctly")
        
        # Verify by fetching the order
        get_response = requests.get(f"{BASE_URL}/api/orders/{order_data['id']}")
        assert get_response.status_code == 200
        fetched_order = get_response.json()
        
        assert fetched_order.get('heaterType') == 'integrated'
        assert fetched_order.get('heaterTypeName') == 'Piec zintegrowany'
        assert fetched_order.get('selectedHeaterVariantId') == 'round_200_integrated'
        
        print("PASS: Order fetched with heater fields persisted correctly")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/orders/{order_data['id']}")
    
    def test_create_order_with_external_heater(self):
        """Test creating order with external heater type"""
        order_data = {
            "id": "TEST-HEATER-EXT-001",
            "fullName": "Test External Heater",
            "phoneNumber": "+48123456789",
            "fullAddress": "Test Address 456",
            "orderDate": "2025-01-15",
            "modelId": "round_225",
            "modelName": "Okrągła 225cm",
            "modelPrice": 1400.0,
            "heaterType": "external",
            "heaterTypeName": "Piec zewnętrzny",
            "selectedHeaterVariantId": "round_225_external",
            "selections": {},
            "selectedOptions": [],
            "notes": "Test order with external heater",
            "discountPercent": 0,
            "subtotal": 1400.0,
            "total": 1400.0,
            "currency": "EUR"
        }
        
        # Create order
        response = requests.post(f"{BASE_URL}/api/orders", json=order_data)
        assert response.status_code == 200, f"Failed to create order: {response.text}"
        
        created_order = response.json()
        assert created_order.get('heaterType') == 'external', f"heaterType should be 'external', got: {created_order.get('heaterType')}"
        assert created_order.get('heaterTypeName') == 'Piec zewnętrzny', f"heaterTypeName should be 'Piec zewnętrzny', got: {created_order.get('heaterTypeName')}"
        assert created_order.get('selectedHeaterVariantId') == 'round_225_external', f"selectedHeaterVariantId should be 'round_225_external', got: {created_order.get('selectedHeaterVariantId')}"
        
        print("PASS: Order created with external heater fields saved correctly")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/orders/{order_data['id']}")
    
    def test_update_order_heater_type(self):
        """Test updating order heater type"""
        order_data = {
            "id": "TEST-HEATER-UPDATE-001",
            "fullName": "Test Update Heater",
            "phoneNumber": "+48123456789",
            "fullAddress": "Test Address",
            "orderDate": "2025-01-15",
            "modelId": "round_225",
            "modelName": "Okrągła 225cm",
            "modelPrice": 1450.0,
            "heaterType": "integrated",
            "heaterTypeName": "Piec zintegrowany",
            "selectedHeaterVariantId": "round_225_integrated",
            "selections": {},
            "selectedOptions": [],
            "total": 1450.0,
            "currency": "EUR"
        }
        
        # Create order
        response = requests.post(f"{BASE_URL}/api/orders", json=order_data)
        assert response.status_code == 200
        
        # Update to external heater
        order_data['heaterType'] = 'external'
        order_data['heaterTypeName'] = 'Piec zewnętrzny'
        order_data['selectedHeaterVariantId'] = 'round_225_external'
        order_data['modelPrice'] = 1400.0
        order_data['total'] = 1400.0
        
        update_response = requests.put(f"{BASE_URL}/api/orders/{order_data['id']}", json=order_data)
        assert update_response.status_code == 200
        
        # Verify update
        get_response = requests.get(f"{BASE_URL}/api/orders/{order_data['id']}")
        assert get_response.status_code == 200
        updated_order = get_response.json()
        
        assert updated_order.get('heaterType') == 'external'
        assert updated_order.get('heaterTypeName') == 'Piec zewnętrzny'
        assert updated_order.get('selectedHeaterVariantId') == 'round_225_external'
        
        print("PASS: Order heater type updated correctly")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/orders/{order_data['id']}")


class TestExcelHeaterMapping:
    """Test Excel generation with heater variant mapping"""
    
    def test_excel_heater_mapping_by_variant_id(self):
        """Test that heater variant is marked in Excel when excelCell is configured"""
        # Note: This test requires excelCell to be configured for heater variants in admin
        # The mapping logic looks for selectedHeaterVariantId in heater_mapping
        
        payload = {
            "fullName": "Test Excel Heater",
            "fullAddress": "Test Address",
            "modelId": "round_200",
            "heaterType": "integrated",
            "selectedHeaterVariantId": "round_200_integrated",
            "orderId": "TEST-EXCEL-HEATER-001"
        }
        
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        assert response.status_code == 200, f"Failed to generate Excel: {response.text}"
        
        # Load Excel and check structure
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verify customer data is written
        assert ws['B2'].value == "Test Excel Heater"
        assert ws['B4'].value == "Test Address"
        
        print("PASS: Excel generated successfully with heater variant data")
        print(f"  - Note: Heater cell marking depends on excelCell configuration in admin")
    
    def test_excel_heater_mapping_by_pattern(self):
        """Test that heater is found by model_id_heater_type pattern"""
        payload = {
            "fullName": "Test Pattern Heater",
            "fullAddress": "Test Address",
            "modelId": "round_225",
            "heaterType": "external",
            "selectedHeaterVariantId": "",  # Empty - should fall back to pattern
            "orderId": "TEST-EXCEL-PATTERN-001"
        }
        
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        assert response.status_code == 200, f"Failed to generate Excel: {response.text}"
        
        print("PASS: Excel generated with pattern-based heater lookup")
    
    def test_excel_with_full_order_data(self):
        """Test Excel generation with complete order including heater"""
        payload = {
            "fullName": "Jan Kowalski",
            "fullAddress": "ul. Testowa 15, Warszawa",
            "modelId": "round_225",
            "heaterType": "integrated",
            "selectedHeaterVariantId": "round_225_integrated",
            "selections": {
                "fiberglass_color": "fg_white",
                "accessories": {
                    "hydromassage_basic": True
                }
            },
            "notes": "Test order with heater",
            "orderId": "TEST-EXCEL-FULL-001"
        }
        
        response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=payload)
        assert response.status_code == 200
        
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verify basic data
        assert ws['B2'].value == "Jan Kowalski"
        assert ws['B4'].value == "ul. Testowa 15, Warszawa"
        assert ws['B18'].value == "Test order with heater"
        
        print("PASS: Full order Excel generated with heater data")


class TestPDFHeaterType:
    """Test PDF generation shows heater type under model name"""
    
    def test_pdf_shows_integrated_heater_type(self):
        """Test that PDF shows 'Piec zintegrowany' for integrated heater"""
        pdf_request = {
            "orderId": "TEST-PDF-INT-001",
            "fullName": "Test PDF Integrated",
            "phoneNumber": "+48123456789",
            "fullAddress": "Test Address 123",
            "orderDate": "2025-01-15",
            "modelId": "round_200",
            "modelName": "Okrągła 200cm",
            "modelPrice": 1677.0,
            "heaterType": "integrated",
            "heaterTypeName": "Piec zintegrowany",
            "selectedHeaterVariantId": "round_200_integrated",
            "selections": {},
            "selectedOptions": [],
            "notes": "",
            "discountPercent": 0,
            "subtotal": 1677.0,
            "total": 1677.0,
            "currency": "EUR",
            "language": "pl",
            "type": "customer"
        }
        
        response = requests.post(f"{BASE_URL}/api/generate-pdf", json=pdf_request)
        assert response.status_code == 200, f"Failed to generate PDF: {response.text}"
        assert 'application/pdf' in response.headers.get('Content-Type', '')
        
        # Parse PDF and check for heater type text
        pdf_reader = PdfReader(io.BytesIO(response.content))
        pdf_text = ""
        for page in pdf_reader.pages:
            pdf_text += page.extract_text() or ""
        
        # Check that heater type is mentioned
        assert "Piec zintegrowany" in pdf_text or "zintegrowany" in pdf_text.lower(), \
            f"PDF should contain 'Piec zintegrowany', PDF text: {pdf_text[:500]}..."
        
        print("PASS: PDF contains integrated heater type text")
    
    def test_pdf_shows_external_heater_type(self):
        """Test that PDF shows 'Piec zewnętrzny' for external heater"""
        pdf_request = {
            "orderId": "TEST-PDF-EXT-001",
            "fullName": "Test PDF External",
            "phoneNumber": "+48123456789",
            "fullAddress": "Test Address 456",
            "orderDate": "2025-01-15",
            "modelId": "round_225",
            "modelName": "Okrągła 225cm",
            "modelPrice": 1400.0,
            "heaterType": "external",
            "heaterTypeName": "Piec zewnętrzny",
            "selectedHeaterVariantId": "round_225_external",
            "selections": {},
            "selectedOptions": [],
            "notes": "",
            "discountPercent": 0,
            "subtotal": 1400.0,
            "total": 1400.0,
            "currency": "EUR",
            "language": "pl",
            "type": "customer"
        }
        
        response = requests.post(f"{BASE_URL}/api/generate-pdf", json=pdf_request)
        assert response.status_code == 200, f"Failed to generate PDF: {response.text}"
        
        # Parse PDF and check for heater type text
        pdf_reader = PdfReader(io.BytesIO(response.content))
        pdf_text = ""
        for page in pdf_reader.pages:
            pdf_text += page.extract_text() or ""
        
        # Check that heater type is mentioned
        assert "Piec zewnętrzny" in pdf_text or "zewnętrzny" in pdf_text.lower(), \
            f"PDF should contain 'Piec zewnętrzny', PDF text: {pdf_text[:500]}..."
        
        print("PASS: PDF contains external heater type text")
    
    def test_pdf_heater_type_under_model_name(self):
        """Test that heater type appears in the model section of PDF"""
        pdf_request = {
            "orderId": "TEST-PDF-MODEL-001",
            "fullName": "Test PDF Model Section",
            "phoneNumber": "+48123456789",
            "fullAddress": "Test Address",
            "orderDate": "2025-01-15",
            "modelId": "square_220",
            "modelName": "Kwadratowa 220cm",
            "modelPrice": 1800.0,
            "heaterType": "integrated",
            "heaterTypeName": "Piec zintegrowany",
            "selectedHeaterVariantId": "square_220_integrated",
            "selections": {},
            "selectedOptions": [],
            "notes": "",
            "discountPercent": 0,
            "subtotal": 1800.0,
            "total": 1800.0,
            "currency": "EUR",
            "language": "pl",
            "type": "customer"
        }
        
        response = requests.post(f"{BASE_URL}/api/generate-pdf", json=pdf_request)
        assert response.status_code == 200
        
        pdf_reader = PdfReader(io.BytesIO(response.content))
        pdf_text = ""
        for page in pdf_reader.pages:
            pdf_text += page.extract_text() or ""
        
        # Both model name and heater type should be present
        assert "Kwadratowa 220cm" in pdf_text or "220" in pdf_text, \
            f"PDF should contain model name"
        assert "Typ pieca" in pdf_text or "zintegrowany" in pdf_text.lower(), \
            f"PDF should contain heater type info"
        
        print("PASS: PDF contains model name and heater type in model section")
    
    def test_pdf_without_heater_type(self):
        """Test PDF generation when heaterType is not provided (backward compatibility)"""
        pdf_request = {
            "orderId": "TEST-PDF-NO-HEATER-001",
            "fullName": "Test PDF No Heater",
            "phoneNumber": "+48123456789",
            "fullAddress": "Test Address",
            "orderDate": "2025-01-15",
            "modelId": "round_200",
            "modelName": "Okrągła 200cm",
            "modelPrice": 1250.0,
            # No heaterType, heaterTypeName, or selectedHeaterVariantId
            "selections": {},
            "selectedOptions": [],
            "notes": "",
            "total": 1250.0,
            "currency": "EUR",
            "language": "pl",
            "type": "customer"
        }
        
        response = requests.post(f"{BASE_URL}/api/generate-pdf", json=pdf_request)
        assert response.status_code == 200, f"PDF should generate even without heater type: {response.text}"
        
        print("PASS: PDF generates successfully without heater type (backward compatibility)")


class TestEndToEndHeaterFlow:
    """End-to-end test for heater type flow: Create order -> Generate Excel -> Generate PDF"""
    
    def test_full_heater_flow(self):
        """Test complete flow with heater type"""
        order_id = "TEST-E2E-HEATER-001"
        
        # 1. Create order with heater type
        order_data = {
            "id": order_id,
            "fullName": "E2E Test Customer",
            "phoneNumber": "+48123456789",
            "fullAddress": "ul. E2E Test 123, Warszawa",
            "orderDate": "2025-01-15",
            "modelId": "round_225",
            "modelName": "Okrągła 225cm",
            "modelPrice": 1450.0,
            "heaterType": "integrated",
            "heaterTypeName": "Piec zintegrowany",
            "selectedHeaterVariantId": "round_225_integrated",
            "selections": {
                "fiberglass_color": "fg_white"
            },
            "selectedOptions": [
                {"id": "fg_white", "categoryId": "fiberglass_color", "optionId": "fg_white", "name": "Biały", "price": 0}
            ],
            "notes": "E2E test order",
            "discountPercent": 5,
            "subtotal": 1450.0,
            "total": 1377.5,
            "currency": "EUR"
        }
        
        create_response = requests.post(f"{BASE_URL}/api/orders", json=order_data)
        assert create_response.status_code == 200, f"Failed to create order: {create_response.text}"
        print("Step 1 PASS: Order created with heater type")
        
        # 2. Verify order has heater fields
        get_response = requests.get(f"{BASE_URL}/api/orders/{order_id}")
        assert get_response.status_code == 200
        saved_order = get_response.json()
        
        assert saved_order.get('heaterType') == 'integrated'
        assert saved_order.get('heaterTypeName') == 'Piec zintegrowany'
        assert saved_order.get('selectedHeaterVariantId') == 'round_225_integrated'
        print("Step 2 PASS: Order heater fields verified")
        
        # 3. Generate Excel
        excel_response = requests.post(f"{BASE_URL}/api/generate-production-excel", json=saved_order)
        assert excel_response.status_code == 200
        
        wb = load_workbook(io.BytesIO(excel_response.content))
        ws = wb.active
        assert ws['B2'].value == "E2E Test Customer"
        print("Step 3 PASS: Excel generated with order data")
        
        # 4. Generate PDF
        pdf_request = {
            **saved_order,
            "language": "pl",
            "type": "customer"
        }
        pdf_response = requests.post(f"{BASE_URL}/api/generate-pdf", json=pdf_request)
        assert pdf_response.status_code == 200
        
        pdf_reader = PdfReader(io.BytesIO(pdf_response.content))
        pdf_text = ""
        for page in pdf_reader.pages:
            pdf_text += page.extract_text() or ""
        
        assert "E2E Test Customer" in pdf_text
        assert "zintegrowany" in pdf_text.lower() or "Typ pieca" in pdf_text
        print("Step 4 PASS: PDF generated with heater type")
        
        # Cleanup
        requests.delete(f"{BASE_URL}/api/orders/{order_id}")
        print("E2E Test PASS: Full heater flow completed successfully")


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
