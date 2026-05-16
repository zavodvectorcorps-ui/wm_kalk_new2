"""Excel/CSV export & import for sauna prices (base + dealer overrides).

Schema (single sheet "Prices") — flat hierarchical table:

  type            id              parentId      category    name      price   costPrice   description   isActive   imageUrl   dealerPrice
  model           sauna_k_235x200                            Kwadro    14200   8000        ...           TRUE       https://   12500
  model_variant   standard        sauna_k_235x200            Standard   0      0           ...           TRUE       https://
  option          piec_9kw        piece         piece       Piec 9kW   2600    1300        ...           TRUE       https://   2400
  option_variant  v1              piec_9kw      piece       Premium    3200    1600        ...           TRUE       https://

Row meaning by `type`:
  - "model"          -> parentId blank; price = basePrice
  - "model_variant"  -> parentId = modelId; price = variant.price
  - "option"         -> parentId = categoryId; price = option.price
  - "option_variant" -> parentId = optionId;  price = variant.price

`dealerPrice` column is only present when exporting a specific dealer; on import it
populates the matching dealer_price_override row. Empty cell = leave override
untouched (do NOT delete). To wipe an override, set price to a number you control
elsewhere (UI).
"""
from __future__ import annotations

import io
import csv
import logging
from typing import Any
from copy import deepcopy

import pandas as pd  # type: ignore
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore

logger = logging.getLogger(__name__)


HEADERS_BASE = [
    "type", "id", "parentId", "category", "name",
    "price", "costPrice", "description", "isActive", "imageUrl",
]


def _truthy(v) -> bool:
    if v is None:
        return True
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("", "true", "1", "yes", "y", "tak", "да"):
        return True
    if s in ("false", "0", "no", "n", "nie", "нет"):
        return False
    return True


def _int_or_none(v):
    """Parse value to int; returns None for blank/NaN values."""
    if v is None:
        return None
    if isinstance(v, float):
        # NaN check
        if v != v:
            return None
        return int(v)
    if isinstance(v, int):
        return v
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return None
    try:
        return int(float(s.replace(",", ".").replace(" ", "")))
    except (ValueError, TypeError):
        return None


def _str_or_empty(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:  # NaN
        return ""
    return str(v).strip()


# ----------------------------------------------------------------------------
# EXPORT
# ----------------------------------------------------------------------------

def _build_rows(prices: dict, overrides_by_key: dict[tuple, int] | None = None,
                include_dealer_price: bool = False) -> list[dict]:
    """Flatten prices doc into a list of row dicts following HEADERS_BASE (+ dealerPrice)."""
    rows: list[dict] = []

    def _dealer_for(key: tuple):
        if not include_dealer_price:
            return ""
        if overrides_by_key is None:
            return ""
        v = overrides_by_key.get(key)
        return "" if v is None else v

    # Models + model_variants
    for m in prices.get("models", []) or []:
        rows.append({
            "type": "model",
            "id": m.get("id", ""),
            "parentId": "",
            "category": "",
            "name": m.get("name", ""),
            "price": int(m.get("basePrice") or 0),
            "costPrice": int(m.get("costPrice") or 0),
            "description": m.get("hint", "") or m.get("websiteDescriptionPl", "") or "",
            "isActive": "TRUE" if m.get("active", True) else "FALSE",
            "imageUrl": m.get("imageUrl", "") or "",
            "dealerPrice": _dealer_for(("model", m.get("id"), None, None, None)),
        })
        for v in m.get("variants", []) or []:
            rows.append({
                "type": "model_variant",
                "id": v.get("id", ""),
                "parentId": m.get("id", ""),
                "category": "",
                "name": v.get("name", "") or v.get("namePl", ""),
                "price": int(v.get("price") or 0),
                "costPrice": int(v.get("costPrice") or 0),
                "description": v.get("hint", "") or v.get("hintPl", "") or "",
                "isActive": "TRUE",
                "imageUrl": v.get("imageUrl", "") or "",
                "dealerPrice": _dealer_for(
                    ("model_variant", m.get("id"), v.get("id"), None, None)
                ),
            })

    # Options (flat list + inside categories)
    def _emit_options(opts: list, category_id: str):
        for opt in opts or []:
            rows.append({
                "type": "option",
                "id": opt.get("id", ""),
                "parentId": category_id,
                "category": category_id,
                "name": opt.get("name", ""),
                "price": int(opt.get("price") or 0),
                "costPrice": int(opt.get("costPrice") or 0),
                "description": opt.get("hint", "") or "",
                "isActive": "TRUE" if opt.get("active", True) else "FALSE",
                "imageUrl": opt.get("imageUrl", "") or "",
                "dealerPrice": _dealer_for(
                    ("option", None, None, opt.get("id"), None)
                ),
            })
            for ov in opt.get("variants", []) or []:
                rows.append({
                    "type": "option_variant",
                    "id": ov.get("id", ""),
                    "parentId": opt.get("id", ""),
                    "category": category_id,
                    "name": ov.get("name", "") or ov.get("namePl", ""),
                    "price": int(ov.get("price") or 0),
                    "costPrice": int(ov.get("costPrice") or 0),
                    "description": "",
                    "isActive": "TRUE",
                    "imageUrl": ov.get("imageUrl", "") or "",
                    "dealerPrice": _dealer_for(
                        ("option_variant", None, None, opt.get("id"), ov.get("id"))
                    ),
                })

    _emit_options(prices.get("options", []) or [], "")
    for cat in prices.get("categories", []) or []:
        _emit_options(cat.get("options", []) or [], cat.get("id", ""))

    return rows


def build_overrides_lookup(overrides: list[dict]) -> dict[tuple, int]:
    out: dict[tuple, int] = {}
    for o in overrides:
        kind = o.get("kind")
        key = (kind, o.get("modelId"), o.get("variantId"),
               o.get("optionId"), o.get("optionVariantId"))
        try:
            out[key] = int(o.get("price") or 0)
        except (TypeError, ValueError):
            continue
    return out


def export_xlsx(prices: dict, overrides: list[dict] | None = None) -> bytes:
    """Export prices to a single-sheet XLSX. Includes dealerPrice column when overrides supplied."""
    include_dealer = overrides is not None
    headers = list(HEADERS_BASE) + (["dealerPrice"] if include_dealer else [])
    lookup = build_overrides_lookup(overrides or []) if include_dealer else None
    rows = _build_rows(prices, lookup, include_dealer)

    wb = Workbook()
    ws = wb.active
    ws.title = "Prices"

    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="left", vertical="center")

    type_fills = {
        "model": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "model_variant": PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid"),
        "option": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "option_variant": PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),
    }

    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        fill = type_fills.get(r["type"])
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    # Column widths
    widths = {
        "type": 16, "id": 32, "parentId": 32, "category": 16, "name": 50,
        "price": 12, "costPrice": 12, "description": 50, "isActive": 10,
        "imageUrl": 50, "dealerPrice": 14,
    }
    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 18)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(prices: dict, overrides: list[dict] | None = None) -> bytes:
    include_dealer = overrides is not None
    headers = list(HEADERS_BASE) + (["dealerPrice"] if include_dealer else [])
    lookup = build_overrides_lookup(overrides or []) if include_dealer else None
    rows = _build_rows(prices, lookup, include_dealer)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({h: r.get(h, "") for h in headers})
    return buf.getvalue().encode("utf-8-sig")


# ----------------------------------------------------------------------------
# IMPORT — parsing
# ----------------------------------------------------------------------------

def parse_file(filename: str, content: bytes) -> list[dict]:
    """Parse uploaded XLSX or CSV into a list of normalized row dicts.

    Returns dicts with raw string values. Validation happens in diff().
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        try:
            df = pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False, encoding="utf-8-sig")
        except UnicodeDecodeError:
            df = pd.read_csv(io.BytesIO(content), dtype=str, keep_default_na=False, encoding="cp1251")
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(io.BytesIO(content), sheet_name=0, dtype=str, keep_default_na=False)
    else:
        raise ValueError("Unsupported file format. Use .xlsx or .csv")

    # Normalize column names
    df.columns = [str(c).strip() for c in df.columns]
    rows = df.to_dict(orient="records")
    return rows


# ----------------------------------------------------------------------------
# DIFF (dry-run)
# ----------------------------------------------------------------------------

DIFF_FIELDS_MODEL = ["price", "costPrice", "description", "isActive", "imageUrl", "name"]
DIFF_FIELDS_OPTION = ["price", "costPrice", "description", "isActive", "imageUrl", "name"]


def _find_model(prices: dict, mid: str):
    for m in prices.get("models", []) or []:
        if m.get("id") == mid:
            return m
    return None


def _find_option_in_cat(cat: dict, oid: str):
    for o in cat.get("options", []) or []:
        if o.get("id") == oid:
            return o
    return None


def _find_option_anywhere(prices: dict, oid: str):
    for o in prices.get("options", []) or []:
        if o.get("id") == oid:
            return o, None
    for cat in prices.get("categories", []) or []:
        o = _find_option_in_cat(cat, oid)
        if o:
            return o, cat
    return None, None


def diff_rows(prices_doc: dict, parsed_rows: list[dict],
              overrides_by_key: dict[tuple, int] | None = None,
              include_dealer_price: bool = False) -> dict:
    """Compare parsed rows to current prices doc; return diff structure.

    Returns:
      {
        "summary": {"added": N, "modified": N, "unchanged": N, "errors": N, "overrides_changed": N},
        "rows": [{"type", "id", "parentId", "name", "status", "diff", "error"}, ...]
      }
    """
    out_rows: list[dict] = []
    summary = {"added": 0, "modified": 0, "unchanged": 0, "errors": 0, "overrides_changed": 0, "marginAlerts": 0}

    # Threshold below which we flag the row as "low margin" (visual warning).
    MARGIN_ALERT_PCT = 15.0

    seen_keys = set()

    for idx, raw in enumerate(parsed_rows):
        row_type = _str_or_empty(raw.get("type")).lower()
        rid = _str_or_empty(raw.get("id"))
        parent_id = _str_or_empty(raw.get("parentId"))
        name = _str_or_empty(raw.get("name"))
        price = _int_or_none(raw.get("price"))
        cost_price = _int_or_none(raw.get("costPrice"))
        description = _str_or_empty(raw.get("description"))
        is_active_raw = raw.get("isActive")
        image_url = _str_or_empty(raw.get("imageUrl"))
        dealer_price_raw = raw.get("dealerPrice") if include_dealer_price else None
        dealer_price = _int_or_none(dealer_price_raw)

        # Skip empty rows
        if not row_type and not rid and not name:
            continue

        result: dict[str, Any] = {
            "rowNumber": idx + 2,  # +2 = 1-indexed header + data starts at 2
            "type": row_type,
            "id": rid,
            "parentId": parent_id,
            "name": name,
            "status": "unchanged",
            "diff": {},
            "error": "",
        }

        # Validate type
        if row_type not in ("model", "model_variant", "option", "option_variant"):
            result["status"] = "error"
            result["error"] = f"Unknown type: '{row_type}'"
            summary["errors"] += 1
            out_rows.append(result)
            continue

        # If id is blank => treat as added (new id generated on commit)
        is_new_id = not rid

        # Lookup existing entity
        existing = None
        if not is_new_id:
            if row_type == "model":
                existing = _find_model(prices_doc, rid)
            elif row_type == "model_variant":
                m = _find_model(prices_doc, parent_id)
                if not m:
                    result["status"] = "error"
                    result["error"] = f"Parent model '{parent_id}' not found"
                    summary["errors"] += 1
                    out_rows.append(result)
                    continue
                existing = next((v for v in (m.get("variants") or []) if v.get("id") == rid), None)
            elif row_type == "option":
                existing, _ = _find_option_anywhere(prices_doc, rid)
            elif row_type == "option_variant":
                opt, _ = _find_option_anywhere(prices_doc, parent_id) if parent_id else (None, None)
                if not opt:
                    result["status"] = "error"
                    result["error"] = f"Parent option '{parent_id}' not found"
                    summary["errors"] += 1
                    out_rows.append(result)
                    continue
                existing = next((v for v in (opt.get("variants") or []) if v.get("id") == rid), None)

        # Diff
        if existing is None:
            # added
            result["status"] = "added"
            new_vals = {
                "name": name,
                "price": price or 0,
                "costPrice": cost_price or 0,
                "description": description,
                "isActive": _truthy(is_active_raw),
                "imageUrl": image_url,
            }
            result["diff"] = {k: {"old": None, "new": v} for k, v in new_vals.items()}
            summary["added"] += 1
        else:
            # Compare fields
            old_price = int((existing.get("basePrice") if row_type == "model" else existing.get("price")) or 0)
            old_cost = int(existing.get("costPrice") or 0)
            old_desc = existing.get("hint", "") or existing.get("websiteDescriptionPl", "") or ""
            old_active_field = existing.get("active", True) if row_type == "model" else existing.get("active", True)
            old_image = existing.get("imageUrl", "") or ""
            old_name = existing.get("name", "") or ""

            new_active = _truthy(is_active_raw) if is_active_raw is not None and str(is_active_raw).strip() != "" else old_active_field

            diff = {}
            if price is not None and price != old_price:
                diff["price"] = {"old": old_price, "new": price}
            if cost_price is not None and cost_price != old_cost:
                diff["costPrice"] = {"old": old_cost, "new": cost_price}
            if description and description != old_desc:
                diff["description"] = {"old": old_desc, "new": description}
            if image_url and image_url != old_image:
                diff["imageUrl"] = {"old": old_image, "new": image_url}
            if name and name != old_name:
                diff["name"] = {"old": old_name, "new": name}
            if new_active != old_active_field:
                diff["isActive"] = {"old": old_active_field, "new": new_active}

            if diff:
                result["status"] = "modified"
                result["diff"] = diff
                summary["modified"] += 1
            else:
                summary["unchanged"] += 1

        # Dealer-price diff (only if include_dealer_price)
        if include_dealer_price and overrides_by_key is not None and dealer_price is not None:
            if row_type == "model":
                ov_key = ("model", rid, None, None, None)
            elif row_type == "model_variant":
                ov_key = ("model_variant", parent_id, rid, None, None)
            elif row_type == "option":
                ov_key = ("option", None, None, rid, None)
            else:
                ov_key = ("option_variant", None, None, parent_id, rid)
            old_dealer = overrides_by_key.get(ov_key)
            if old_dealer != dealer_price:
                result["diff"]["dealerPrice"] = {"old": old_dealer, "new": dealer_price}
                if result["status"] == "unchanged":
                    result["status"] = "modified"
                    summary["modified"] += 1
                    summary["unchanged"] -= 1
                summary["overrides_changed"] += 1

        # Track keys
        key = (row_type, rid or f"_new_{idx}", parent_id)
        if key in seen_keys:
            result["status"] = "error"
            result["error"] = "Duplicate row (same type+id+parentId)"
            summary["errors"] += 1
            # adjust counters: if we already incremented added/modified/unchanged above, revert
        else:
            seen_keys.add(key)

        # ---- Margin computation (price - costPrice) ----
        if existing is None:
            old_price_for_margin = None
            old_cost_for_margin = None
        else:
            old_price_for_margin = int((existing.get("basePrice") if row_type == "model" else existing.get("price")) or 0)
            old_cost_for_margin = int(existing.get("costPrice") or 0)
        new_price_for_margin = price if price is not None else old_price_for_margin
        new_cost_for_margin = cost_price if cost_price is not None else old_cost_for_margin

        def _margin(p, c):
            if p is None or c is None:
                return None
            return int(p) - int(c)

        def _margin_pct(p, c):
            if p is None or c is None or int(p) <= 0:
                return None
            return round((int(p) - int(c)) * 100.0 / int(p), 1)

        margin_old = _margin(old_price_for_margin, old_cost_for_margin)
        margin_new = _margin(new_price_for_margin, new_cost_for_margin)
        margin_pct_old = _margin_pct(old_price_for_margin, old_cost_for_margin)
        margin_pct_new = _margin_pct(new_price_for_margin, new_cost_for_margin)
        result["margin"] = {
            "oldAmount": margin_old,
            "newAmount": margin_new,
            "oldPct": margin_pct_old,
            "newPct": margin_pct_new,
            "delta": (margin_new - margin_old) if (margin_old is not None and margin_new is not None) else None,
        }

        # Flag low margin (only on the new values, when result has a real price)
        if (
            margin_pct_new is not None
            and margin_pct_new < MARGIN_ALERT_PCT
            and result["status"] in ("added", "modified")
            and (new_price_for_margin or 0) > 0
        ):
            result["lowMargin"] = True
            result["marginThreshold"] = MARGIN_ALERT_PCT
            summary["marginAlerts"] += 1

        out_rows.append(result)

    return {"summary": summary, "rows": out_rows}


# ----------------------------------------------------------------------------
# SNAPSHOT DIFF — compare two prices docs (used by history-entry diff view)
# ----------------------------------------------------------------------------

def _row_key(r: dict) -> tuple:
    """Unique key for a row from `_build_rows`: (type, id, parentId)."""
    return (r.get("type"), r.get("id"), r.get("parentId") or "")


def snapshot_diff(before_prices: dict | None, after_prices: dict | None,
                  before_overrides: list[dict] | None = None,
                  after_overrides: list[dict] | None = None,
                  include_dealer_price: bool = False) -> dict:
    """Diff two snapshots of the prices document.

    Returns the same shape as `diff_rows`:
      {"summary": {added, modified, removed, unchanged, marginAlerts}, "rows": [...]}

    Row statuses:
      - "added"     — present in AFTER, not in BEFORE
      - "modified"  — present in both, at least one field changed
      - "removed"   — present in BEFORE, not in AFTER
      - "unchanged" — present in both, identical
    """
    before_prices = before_prices or {}
    after_prices = after_prices or {}

    before_lookup = build_overrides_lookup(before_overrides or []) if include_dealer_price else None
    after_lookup = build_overrides_lookup(after_overrides or []) if include_dealer_price else None

    before_rows = {_row_key(r): r for r in _build_rows(before_prices, before_lookup, include_dealer_price)}
    after_rows = {_row_key(r): r for r in _build_rows(after_prices, after_lookup, include_dealer_price)}

    all_keys = list(dict.fromkeys(list(before_rows.keys()) + list(after_rows.keys())))

    summary = {"added": 0, "modified": 0, "removed": 0, "unchanged": 0, "marginAlerts": 0}
    MARGIN_ALERT_PCT = 15.0
    out: list[dict] = []

    fields = ("name", "price", "costPrice", "description", "isActive", "imageUrl")
    if include_dealer_price:
        fields = fields + ("dealerPrice",)

    for k in all_keys:
        b = before_rows.get(k)
        a = after_rows.get(k)
        row_type, rid, parent_id = k

        # Pick a row to display the entity meta from (prefer "after" — current state)
        meta_src = a or b or {}
        result: dict[str, Any] = {
            "type": row_type,
            "id": rid,
            "parentId": parent_id,
            "name": meta_src.get("name", ""),
            "status": "unchanged",
            "diff": {},
            "error": "",
            "margin": {},
        }

        if b is None and a is not None:
            result["status"] = "added"
            result["diff"] = {f: {"old": None, "new": a.get(f)} for f in fields if a.get(f) not in (None, "", 0)}
            summary["added"] += 1
        elif a is None and b is not None:
            result["status"] = "removed"
            result["diff"] = {f: {"old": b.get(f), "new": None} for f in fields if b.get(f) not in (None, "", 0)}
            summary["removed"] += 1
        else:
            # Both present — compute field diffs
            diff = {}
            for f in fields:
                bv, av = b.get(f), a.get(f)
                # Treat empty/None as equal
                if (bv or None) != (av or None):
                    diff[f] = {"old": bv, "new": av}
            if diff:
                result["status"] = "modified"
                result["diff"] = diff
                summary["modified"] += 1
            else:
                summary["unchanged"] += 1

        # Margin
        def _to_int(v):
            try:
                return int(v) if v is not None and v != "" else None
            except (TypeError, ValueError):
                return None

        old_p = _to_int(b.get("price")) if b else None
        old_c = _to_int(b.get("costPrice")) if b else None
        new_p = _to_int(a.get("price")) if a else None
        new_c = _to_int(a.get("costPrice")) if a else None

        def _margin(p, c):
            return None if p is None or c is None else int(p) - int(c)

        def _margin_pct(p, c):
            if p is None or c is None or int(p) <= 0:
                return None
            return round((int(p) - int(c)) * 100.0 / int(p), 1)

        m_old = _margin(old_p, old_c)
        m_new = _margin(new_p, new_c)
        mp_old = _margin_pct(old_p, old_c)
        mp_new = _margin_pct(new_p, new_c)
        result["margin"] = {
            "oldAmount": m_old,
            "newAmount": m_new,
            "oldPct": mp_old,
            "newPct": mp_new,
            "delta": (m_new - m_old) if (m_old is not None and m_new is not None) else None,
        }
        if mp_new is not None and mp_new < MARGIN_ALERT_PCT and result["status"] in ("added", "modified") and (new_p or 0) > 0:
            result["lowMargin"] = True
            result["marginThreshold"] = MARGIN_ALERT_PCT
            summary["marginAlerts"] += 1

        out.append(result)

    return {"summary": summary, "rows": out}


# ----------------------------------------------------------------------------
# COMMIT — apply parsed rows
# ----------------------------------------------------------------------------
def apply_rows(prices_doc: dict, parsed_rows: list[dict],
               include_dealer_price: bool = False) -> tuple[dict, list[dict], dict]:
    """Return (updated_prices_doc, dealer_override_changes, summary).

    dealer_override_changes is a list of {kind, modelId, variantId, optionId, optionVariantId, price}
    to be upserted. Empty/missing dealerPrice is ignored.
    """
    import uuid

    doc = deepcopy(prices_doc)
    summary = {"added": 0, "modified": 0, "unchanged": 0, "errors": 0, "overrides_changed": 0}
    overrides_to_write: list[dict] = []

    for raw in parsed_rows:
        row_type = _str_or_empty(raw.get("type")).lower()
        rid = _str_or_empty(raw.get("id"))
        parent_id = _str_or_empty(raw.get("parentId"))
        name = _str_or_empty(raw.get("name"))
        price = _int_or_none(raw.get("price"))
        cost_price = _int_or_none(raw.get("costPrice"))
        description = _str_or_empty(raw.get("description"))
        is_active_raw = raw.get("isActive")
        image_url = _str_or_empty(raw.get("imageUrl"))
        dealer_price = _int_or_none(raw.get("dealerPrice")) if include_dealer_price else None

        if not row_type and not rid and not name:
            continue
        if row_type not in ("model", "model_variant", "option", "option_variant"):
            summary["errors"] += 1
            continue

        new_id = rid or str(uuid.uuid4())[:8]

        if row_type == "model":
            models = doc.setdefault("models", [])
            existing = next((m for m in models if m.get("id") == rid), None) if rid else None
            if existing is None:
                models.append({
                    "id": new_id,
                    "name": name,
                    "basePrice": price or 0,
                    "costPrice": cost_price or 0,
                    "imageUrl": image_url,
                    "hint": description,
                    "active": _truthy(is_active_raw),
                    "sortOrder": len(models) + 1,
                    "variants": [],
                })
                summary["added"] += 1
                rid = new_id
            else:
                changed = False
                if name and existing.get("name") != name:
                    existing["name"] = name; changed = True
                if price is not None and int(existing.get("basePrice") or 0) != price:
                    existing["basePrice"] = price; changed = True
                if cost_price is not None and int(existing.get("costPrice") or 0) != cost_price:
                    existing["costPrice"] = cost_price; changed = True
                if description and existing.get("hint", "") != description:
                    existing["hint"] = description; changed = True
                if image_url and existing.get("imageUrl", "") != image_url:
                    existing["imageUrl"] = image_url; changed = True
                if is_active_raw is not None and str(is_active_raw).strip() != "":
                    new_a = _truthy(is_active_raw)
                    if existing.get("active", True) != new_a:
                        existing["active"] = new_a; changed = True
                if changed:
                    summary["modified"] += 1
                else:
                    summary["unchanged"] += 1

        elif row_type == "model_variant":
            m = next((x for x in doc.get("models", []) if x.get("id") == parent_id), None)
            if not m:
                summary["errors"] += 1
                continue
            variants = m.setdefault("variants", [])
            existing = next((v for v in variants if v.get("id") == rid), None) if rid else None
            if existing is None:
                variants.append({
                    "id": new_id,
                    "name": name,
                    "price": price or 0,
                    "costPrice": cost_price or 0,
                    "imageUrl": image_url,
                    "hint": description,
                })
                summary["added"] += 1
                rid = new_id
            else:
                changed = False
                if name and existing.get("name") != name:
                    existing["name"] = name; changed = True
                if price is not None and int(existing.get("price") or 0) != price:
                    existing["price"] = price; changed = True
                if cost_price is not None and int(existing.get("costPrice") or 0) != cost_price:
                    existing["costPrice"] = cost_price; changed = True
                if description and (existing.get("hint", "") or "") != description:
                    existing["hint"] = description; changed = True
                if image_url and (existing.get("imageUrl", "") or "") != image_url:
                    existing["imageUrl"] = image_url; changed = True
                if changed:
                    summary["modified"] += 1
                else:
                    summary["unchanged"] += 1

        elif row_type == "option":
            # Find existing option anywhere
            target_opt = None
            target_cat = None
            if rid:
                for o in doc.get("options", []) or []:
                    if o.get("id") == rid:
                        target_opt = o; target_cat = None; break
                if target_opt is None:
                    for cat in doc.get("categories", []) or []:
                        for o in cat.get("options", []) or []:
                            if o.get("id") == rid:
                                target_opt = o; target_cat = cat; break
                        if target_opt is not None:
                            break

            if target_opt is None:
                # add new — into category by parentId, or into flat options
                new_opt = {
                    "id": new_id,
                    "name": name,
                    "price": price or 0,
                    "costPrice": cost_price or 0,
                    "imageUrl": image_url,
                    "hint": description,
                    "inputType": "radio",
                    "sortOrder": 99,
                    "active": _truthy(is_active_raw),
                }
                if parent_id:
                    cat = next((c for c in doc.get("categories", []) or [] if c.get("id") == parent_id), None)
                    if cat is None:
                        summary["errors"] += 1
                        continue
                    cat.setdefault("options", []).append(new_opt)
                else:
                    doc.setdefault("options", []).append(new_opt)
                summary["added"] += 1
                rid = new_id
            else:
                changed = False
                if name and target_opt.get("name") != name:
                    target_opt["name"] = name; changed = True
                if price is not None and int(target_opt.get("price") or 0) != price:
                    target_opt["price"] = price; changed = True
                if cost_price is not None and int(target_opt.get("costPrice") or 0) != cost_price:
                    target_opt["costPrice"] = cost_price; changed = True
                if description and (target_opt.get("hint", "") or "") != description:
                    target_opt["hint"] = description; changed = True
                if image_url and (target_opt.get("imageUrl", "") or "") != image_url:
                    target_opt["imageUrl"] = image_url; changed = True
                if is_active_raw is not None and str(is_active_raw).strip() != "":
                    new_a = _truthy(is_active_raw)
                    if target_opt.get("active", True) != new_a:
                        target_opt["active"] = new_a; changed = True
                if changed:
                    summary["modified"] += 1
                else:
                    summary["unchanged"] += 1

        elif row_type == "option_variant":
            # Find parent option
            target_opt = None
            for o in doc.get("options", []) or []:
                if o.get("id") == parent_id:
                    target_opt = o; break
            if target_opt is None:
                for cat in doc.get("categories", []) or []:
                    for o in cat.get("options", []) or []:
                        if o.get("id") == parent_id:
                            target_opt = o; break
                    if target_opt:
                        break
            if target_opt is None:
                summary["errors"] += 1
                continue
            variants = target_opt.setdefault("variants", [])
            existing = next((v for v in variants if v.get("id") == rid), None) if rid else None
            if existing is None:
                variants.append({
                    "id": new_id,
                    "name": name,
                    "price": price or 0,
                    "costPrice": cost_price or 0,
                    "imageUrl": image_url,
                })
                summary["added"] += 1
                rid = new_id
            else:
                changed = False
                if name and existing.get("name") != name:
                    existing["name"] = name; changed = True
                if price is not None and int(existing.get("price") or 0) != price:
                    existing["price"] = price; changed = True
                if cost_price is not None and int(existing.get("costPrice") or 0) != cost_price:
                    existing["costPrice"] = cost_price; changed = True
                if image_url and (existing.get("imageUrl", "") or "") != image_url:
                    existing["imageUrl"] = image_url; changed = True
                if changed:
                    summary["modified"] += 1
                else:
                    summary["unchanged"] += 1

        # Build dealer override record (regardless of base-price changes)
        if include_dealer_price and dealer_price is not None and rid:
            if row_type == "model":
                overrides_to_write.append({
                    "kind": "model", "modelId": rid, "price": dealer_price,
                })
            elif row_type == "model_variant":
                overrides_to_write.append({
                    "kind": "model_variant", "modelId": parent_id, "variantId": rid, "price": dealer_price,
                })
            elif row_type == "option":
                overrides_to_write.append({
                    "kind": "option", "optionId": rid, "price": dealer_price,
                })
            elif row_type == "option_variant":
                overrides_to_write.append({
                    "kind": "option_variant", "optionId": parent_id, "optionVariantId": rid, "price": dealer_price,
                })
            summary["overrides_changed"] += 1

    return doc, overrides_to_write, summary
