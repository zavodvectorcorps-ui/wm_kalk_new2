"""Excel export/import for sauna_components + sauna_tech_cards.

Layout of the workbook (single file, two sheets):

  Sheet `Components`
    columns: id | name | category | unit | unitPrice | supplier | note |
             stockCurrent | stockMin | isActive

  Sheet `TechCards`
    Each row = ONE BOM item in ONE card.
    Repeated card-level fields (laborCost, overheadPct, etc.) appear on each
    row of the same card; on import the LAST non-empty value wins (per card).
    columns: cardId? | scope | modelId | variantId | optionId | optionVariantId |
             componentId | componentName? | qty | itemNote? |
             laborCost | overheadPct | manualAdjustment | syncToCostPrice | cardNote
"""
from __future__ import annotations

import io
import uuid
from copy import deepcopy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


COMP_HEADERS = [
    "id", "name", "category", "unit", "unitPrice", "supplier", "note",
    "stockCurrent", "stockMin", "isActive",
]

CARD_HEADERS = [
    "cardId", "scope", "modelId", "variantId", "optionId", "optionVariantId",
    "componentId", "componentName", "qty", "itemNote",
    "laborCost", "overheadPct", "manualAdjustment", "syncToCostPrice", "cardNote",
]

VALID_SCOPES = ("model", "variant", "option", "option_variant")


# --------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------
def _s(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _bool(v) -> bool:
    if isinstance(v, bool):
        return v
    s = _s(v).lower()
    return s in ("1", "true", "yes", "y", "+", "да", "истина")


# --------------------------------------------------------------------
# EXPORT
# --------------------------------------------------------------------
def export_xlsx(components: list[dict], tech_cards: list[dict]) -> bytes:
    wb = Workbook()

    # ----- Components -----
    ws_c = wb.active
    ws_c.title = "Components"
    _write_header(ws_c, COMP_HEADERS)
    for c in components:
        ws_c.append([
            _s(c.get("id")),
            _s(c.get("name")),
            _s(c.get("category") or "other"),
            _s(c.get("unit") or "шт"),
            c.get("unitPrice") or 0,
            _s(c.get("supplier")),
            _s(c.get("note")),
            c.get("stockCurrent") or 0,
            c.get("stockMin") or 0,
            _bool(c.get("isActive", True)),
        ])
    _autosize(ws_c, COMP_HEADERS, {"name": 40, "note": 36, "supplier": 28})

    # ----- TechCards (flatten) -----
    ws_t = wb.create_sheet("TechCards")
    _write_header(ws_t, CARD_HEADERS)
    # Build a fast lookup of component name by id, so the exported sheet is
    # readable even before clicking into individual cards.
    comp_by_id = {c.get("id"): c for c in components}
    for card in tech_cards:
        items = card.get("items") or []
        if not items:
            # Card with no items — still record the card row (so user can fill items)
            ws_t.append([
                _s(card.get("id")),
                _s(card.get("scope")),
                _s(card.get("modelId")),
                _s(card.get("variantId")),
                _s(card.get("optionId")),
                _s(card.get("optionVariantId")),
                "", "", "", "",
                card.get("laborCost") or 0,
                card.get("overheadPct") or 0,
                card.get("manualAdjustment") or 0,
                _bool(card.get("syncToCostPrice", True)),
                _s(card.get("note")),
            ])
            continue
        for idx, it in enumerate(items):
            comp = comp_by_id.get(it.get("componentId")) or {}
            row = [
                _s(card.get("id")),
                _s(card.get("scope")),
                _s(card.get("modelId")),
                _s(card.get("variantId")),
                _s(card.get("optionId")),
                _s(card.get("optionVariantId")),
                _s(it.get("componentId")),
                _s(comp.get("name")),
                it.get("qty") or 0,
                _s(it.get("note")),
            ]
            # Only first row of a card carries card-level fields, to keep the
            # sheet visually readable.
            if idx == 0:
                row += [
                    card.get("laborCost") or 0,
                    card.get("overheadPct") or 0,
                    card.get("manualAdjustment") or 0,
                    _bool(card.get("syncToCostPrice", True)),
                    _s(card.get("note")),
                ]
            else:
                row += ["", "", "", "", ""]
            ws_t.append(row)
    _autosize(ws_t, CARD_HEADERS, {"componentName": 40, "cardNote": 36})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header(ws, headers: list[str]):
    fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for c in ws[1]:
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, headers: list[str], overrides: dict | None = None):
    overrides = overrides or {}
    for idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = overrides.get(h, 18)


# --------------------------------------------------------------------
# IMPORT — parse
# --------------------------------------------------------------------
def parse_xlsx(file_bytes: bytes) -> tuple[list[dict], list[dict], list[dict]]:
    """Return (components, cards, errors).

    components: list of dicts shaped like sauna_components docs (no _id).
    cards:      list of dicts grouped by (cardId|key) with items[].
    errors:     [{sheet, row, col?, message}]
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    errors: list[dict] = []

    components = _parse_components(wb, errors)
    cards = _parse_cards(wb, errors)
    return components, cards, errors


def _parse_components(wb, errors: list[dict]) -> list[dict]:
    if "Components" not in wb.sheetnames:
        return []
    ws = wb["Components"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [_s(c).lower() for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    out: list[dict] = []
    for r_idx, raw in enumerate(rows[1:], start=2):
        if not any(_s(v) for v in raw):
            continue
        name = _s(raw[idx.get("name", -1)]) if "name" in idx else ""
        if not name:
            errors.append({"sheet": "Components", "row": r_idx, "message": "name is required"})
            continue
        unit_price = _num(raw[idx.get("unitprice", -1)]) if "unitprice" in idx else 0
        if unit_price is None:
            unit_price = 0
        stock_cur = _num(raw[idx.get("stockcurrent", -1)]) if "stockcurrent" in idx else 0
        stock_min = _num(raw[idx.get("stockmin", -1)]) if "stockmin" in idx else 0
        out.append({
            "id": _s(raw[idx["id"]]) if "id" in idx and idx["id"] < len(raw) else "",
            "name": name,
            "category": _s(raw[idx["category"]]) if "category" in idx else "other",
            "unit": _s(raw[idx["unit"]]) if "unit" in idx else "шт",
            "unitPrice": float(unit_price or 0),
            "supplier": _s(raw[idx["supplier"]]) if "supplier" in idx else "",
            "note": _s(raw[idx["note"]]) if "note" in idx else "",
            "stockCurrent": float(stock_cur or 0),
            "stockMin": float(stock_min or 0),
            "isActive": _bool(raw[idx["isactive"]]) if "isactive" in idx else True,
        })
    return out


def _parse_cards(wb, errors: list[dict]) -> list[dict]:
    if "TechCards" not in wb.sheetnames:
        return []
    ws = wb["TechCards"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [_s(c).lower() for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}

    def col(raw, name):
        i = idx.get(name.lower())
        if i is None or i >= len(raw):
            return None
        return raw[i]

    cards_by_key: dict[tuple, dict] = {}
    for r_idx, raw in enumerate(rows[1:], start=2):
        if not any(_s(v) for v in raw):
            continue
        scope = _s(col(raw, "scope")).lower()
        if scope not in VALID_SCOPES:
            errors.append({"sheet": "TechCards", "row": r_idx, "message": f"invalid scope {scope!r}"})
            continue
        model_id = _s(col(raw, "modelId"))
        variant_id = _s(col(raw, "variantId"))
        option_id = _s(col(raw, "optionId"))
        option_variant_id = _s(col(raw, "optionVariantId"))
        card_id = _s(col(raw, "cardId"))
        # Group by (cardId or scope-key)
        key = card_id or f"{scope}|{model_id}|{variant_id}|{option_id}|{option_variant_id}"
        card = cards_by_key.get(key)
        if card is None:
            card = {
                "id": card_id or None,
                "scope": scope,
                "modelId": model_id or None,
                "variantId": variant_id or None,
                "optionId": option_id or None,
                "optionVariantId": option_variant_id or None,
                "items": [],
                "laborCost": None,
                "overheadPct": None,
                "manualAdjustment": None,
                "syncToCostPrice": None,
                "note": None,
            }
            cards_by_key[key] = card

        # Card-level fields: last non-empty wins
        for fld, parser in (
            ("laborCost", _num), ("overheadPct", _num),
            ("manualAdjustment", _num), ("cardNote", _s),
        ):
            v = col(raw, fld)
            if v not in (None, ""):
                target_key = "note" if fld == "cardNote" else fld
                card[target_key] = parser(v)
        v = col(raw, "syncToCostPrice")
        if v not in (None, ""):
            card["syncToCostPrice"] = _bool(v)

        # Item row
        comp_id = _s(col(raw, "componentId"))
        qty = _num(col(raw, "qty"))
        if comp_id and qty is not None and qty != 0:
            card["items"].append({
                "componentId": comp_id,
                "qty": float(qty),
                "note": _s(col(raw, "itemNote")) or "",
            })

    return list(cards_by_key.values())


# --------------------------------------------------------------------
# DRY-RUN diff
# --------------------------------------------------------------------
def diff_components(parsed: list[dict], existing: list[dict]) -> dict:
    """Return {add, update, unchanged, byId, byName}."""
    by_id = {c["id"]: c for c in existing if c.get("id")}
    by_name = {c.get("name", "").strip().lower(): c for c in existing if c.get("name")}
    add, update, unchanged = [], [], []

    def _norm_num(v):
        try:
            return float(v or 0)
        except (TypeError, ValueError):
            return 0.0

    def _norm_str(v):
        return (str(v) if v is not None else "").strip()

    NUM_FIELDS = ("unitPrice", "stockCurrent", "stockMin")
    STR_FIELDS = ("name", "category", "unit", "supplier", "note")
    BOOL_FIELDS = ("isActive",)

    for p in parsed:
        # Match by id first; if id is empty, try by name (case-insensitive)
        match = by_id.get(p.get("id"))
        if not match and p.get("name"):
            match = by_name.get(p["name"].strip().lower())
        if not match:
            add.append(p)
            continue
        changed: dict[str, dict] = {}
        for fld in STR_FIELDS:
            if _norm_str(match.get(fld)) != _norm_str(p.get(fld)):
                changed[fld] = {"old": match.get(fld), "new": p.get(fld)}
        for fld in NUM_FIELDS:
            if _norm_num(match.get(fld)) != _norm_num(p.get(fld)):
                changed[fld] = {"old": match.get(fld), "new": p.get(fld)}
        for fld in BOOL_FIELDS:
            if bool(match.get(fld, True)) != bool(p.get(fld, True)):
                changed[fld] = {"old": match.get(fld), "new": p.get(fld)}
        if changed:
            update.append({"id": match["id"], "name": p["name"], "changes": changed})
        else:
            unchanged.append({"id": match["id"], "name": p["name"]})
    return {"add": add, "update": update, "unchanged": unchanged}


def diff_cards(parsed: list[dict], existing: list[dict]) -> dict:
    """Return {add, update, unchanged}."""
    by_key = {
        (c.get("scope"), c.get("modelId") or "", c.get("variantId") or "",
         c.get("optionId") or "", c.get("optionVariantId") or ""): c
        for c in existing
    }
    by_id = {c["id"]: c for c in existing if c.get("id")}

    add, update, unchanged = [], [], []
    for p in parsed:
        key = (
            p["scope"], p.get("modelId") or "", p.get("variantId") or "",
            p.get("optionId") or "", p.get("optionVariantId") or "",
        )
        match = by_id.get(p.get("id")) if p.get("id") else None
        if not match:
            match = by_key.get(key)
        if not match:
            add.append({
                "scope": p["scope"], "modelId": p.get("modelId"),
                "variantId": p.get("variantId"), "optionId": p.get("optionId"),
                "optionVariantId": p.get("optionVariantId"),
                "itemsCount": len(p.get("items") or []),
            })
            continue
        old_items = sorted([(i.get("componentId"), float(i.get("qty") or 0)) for i in (match.get("items") or [])])
        new_items = sorted([(i["componentId"], float(i["qty"])) for i in (p.get("items") or [])])
        changed: dict[str, dict] = {}
        if old_items != new_items:
            changed["items"] = {"old": len(old_items), "new": len(new_items)}
        for fld in ("laborCost", "overheadPct", "manualAdjustment"):
            new = p.get(fld)
            if new is None:
                continue
            if float(match.get(fld) or 0) != float(new):
                changed[fld] = {"old": match.get(fld), "new": new}
        if p.get("syncToCostPrice") is not None and bool(match.get("syncToCostPrice", True)) != bool(p["syncToCostPrice"]):
            changed["syncToCostPrice"] = {"old": match.get("syncToCostPrice"), "new": p["syncToCostPrice"]}
        if changed:
            update.append({"id": match["id"], "key": key, "changes": changed})
        else:
            unchanged.append({"id": match["id"], "key": key})
    return {"add": add, "update": update, "unchanged": unchanged}


# --------------------------------------------------------------------
# COMMIT helpers
# --------------------------------------------------------------------
def merge_component(parsed: dict, existing: dict | None) -> dict:
    """Build a final component doc to upsert."""
    base = deepcopy(existing) if existing else {}
    base["id"] = base.get("id") or parsed.get("id") or str(uuid.uuid4())
    for fld in ("name", "category", "unit", "supplier", "note"):
        if parsed.get(fld) not in (None, ""):
            base[fld] = parsed[fld]
    for fld in ("unitPrice", "stockCurrent", "stockMin"):
        if parsed.get(fld) is not None:
            base[fld] = float(parsed[fld])
    if "isActive" in parsed:
        base["isActive"] = bool(parsed["isActive"])
    return base


def merge_card(parsed: dict, existing: dict | None) -> dict:
    """Build a final tech-card doc (without computed totals — caller recomputes)."""
    base = deepcopy(existing) if existing else {}
    base["id"] = base.get("id") or parsed.get("id") or str(uuid.uuid4())
    base["scope"] = parsed["scope"]
    base["modelId"] = parsed.get("modelId") or ""
    base["variantId"] = parsed.get("variantId") or ""
    base["optionId"] = parsed.get("optionId") or ""
    base["optionVariantId"] = parsed.get("optionVariantId") or ""
    base["items"] = [
        {"id": str(uuid.uuid4()), "componentId": i["componentId"],
         "qty": float(i["qty"]), "note": i.get("note", "")}
        for i in (parsed.get("items") or [])
    ]
    for fld in ("laborCost", "overheadPct", "manualAdjustment"):
        if parsed.get(fld) is not None:
            base[fld] = float(parsed[fld])
    if parsed.get("syncToCostPrice") is not None:
        base["syncToCostPrice"] = bool(parsed["syncToCostPrice"])
    elif "syncToCostPrice" not in base:
        base["syncToCostPrice"] = True
    if parsed.get("note") is not None:
        base["note"] = parsed["note"]
    return base
