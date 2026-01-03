"""Balia calculator routes."""
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import List
from datetime import datetime, timedelta, timezone
from urllib.parse import quote
import io
import os
import logging
import time

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from database import db
from models.balia import PriceData, Order, PDFRequest
from data.balia_defaults import default_balia_prices
from services.telegram_service import notify_new_order

logger = logging.getLogger(__name__)
router = APIRouter(tags=["Balia Calculator"])

# Simple in-memory cache for prices (60 seconds TTL)
_prices_cache = {"data": None, "expires": 0}
CACHE_TTL = 60  # seconds


@router.get("/")
async def root():
    return {"message": "Hot Tub Calculator API"}


@router.get("/prices")
async def get_prices():
    """Get current pricing"""
    prices = await db.prices.find_one({"_id": "default"})
    if not prices:
        await db.prices.insert_one({"_id": "default", **default_balia_prices})
        return default_balia_prices
    
    prices.pop('_id', None)
    
    # Ensure critical fields are arrays
    if not isinstance(prices.get('models'), list):
        prices['models'] = default_balia_prices.get('models', [])
    if not isinstance(prices.get('categories'), list):
        prices['categories'] = default_balia_prices.get('categories', [])
    
    # === MERGE MISSING CATEGORIES FROM DEFAULTS ===
    # Get existing category IDs
    existing_cat_ids = {cat.get('id') for cat in prices.get('categories', [])}
    
    # Add missing categories from defaults (e.g., bowl_material, fiberglass_color, acrylic_color)
    categories_added = False
    for default_cat in default_balia_prices.get('categories', []):
        if default_cat.get('id') not in existing_cat_ids:
            prices['categories'].append(default_cat)
            categories_added = True
            logger.info(f"Added missing category from defaults: {default_cat.get('id')}")
    
    # Sort categories by sortOrder after adding new ones
    if categories_added:
        prices['categories'].sort(key=lambda c: c.get('sortOrder', 999))
        # Save the updated categories to DB
        await db.prices.update_one(
            {"_id": "default"},
            {"$set": {"categories": prices['categories']}}
        )
    
    # Merge hints from defaults if missing in DB data
    default_model_hints = {m['id']: m.get('hint', '') for m in default_balia_prices.get('models', [])}
    default_option_hints = {}
    for cat in default_balia_prices.get('categories', []):
        for opt in cat.get('options', []):
            default_option_hints[opt['id']] = opt.get('hint', '')
    
    # Add hints to models if missing
    for model in prices.get('models', []):
        if not model.get('hint') and model.get('id') in default_model_hints:
            model['hint'] = default_model_hints[model['id']]
    
    # Add hints to options if missing
    for category in prices.get('categories', []):
        for option in category.get('options', []):
            if not option.get('hint') and option.get('id') in default_option_hints:
                option['hint'] = default_option_hints[option['id']]
    
    # Ensure currency fields exist
    prices.setdefault('currency', 'EUR')
    prices.setdefault('currencySymbol', '€')
    
    # === ENSURE HEATER VARIANTS HAVE IDs ===
    variants_updated = False
    for model in prices.get('models', []):
        model_id = model.get('id', '')
        for hv in model.get('heaterVariants', []):
            if not hv.get('id'):
                # Generate ID from model_id + heater type
                hv['id'] = f"{model_id}_{hv.get('type', 'unknown')}"
                variants_updated = True
                logger.info(f"Generated heater variant ID: {hv['id']}")
    
    # Save if variants were updated
    if variants_updated:
        await db.prices.update_one(
            {"_id": "default"},
            {"$set": {"models": prices['models']}}
        )
    
    return prices


@router.post("/prices")
async def update_prices(prices: PriceData):
    """Update pricing"""
    global _prices_cache
    price_dict = prices.model_dump()
    await db.prices.update_one(
        {"_id": "default"},
        {"$set": price_dict},
        upsert=True
    )
    # Invalidate cache
    _prices_cache = {"data": None, "expires": 0}
    return {"message": "Prices updated successfully"}


@router.get("/prices/export")
async def export_prices():
    """Export prices to Excel file"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    prices = await db.prices.find_one({"_id": "default"})
    if not prices:
        raise HTTPException(status_code=404, detail="Prices not found")
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === Sheet 1: Models ===
    ws_models = wb.active
    ws_models.title = "Modele"
    
    # Header
    model_headers = ["ID", "Nazwa (RU)", "Nazwa (PL)", "Typ печi", "Zakup EUR", "Marża %", "Cena PLN", "Kolor HEX"]
    for col, header in enumerate(model_headers, 1):
        cell = ws_models.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for model in prices.get('models', []):
        model_name_ru = model.get('nameRu', model.get('name', ''))
        model_name_pl = model.get('namePl', model.get('name', ''))
        
        for variant in model.get('heaterVariants', []):
            ws_models.cell(row=row, column=1, value=model.get('id', '')).border = thin_border
            ws_models.cell(row=row, column=2, value=model_name_ru).border = thin_border
            ws_models.cell(row=row, column=3, value=model_name_pl).border = thin_border
            ws_models.cell(row=row, column=4, value=variant.get('type', '')).border = thin_border
            ws_models.cell(row=row, column=5, value=variant.get('purchasePriceEur', 0)).border = thin_border
            ws_models.cell(row=row, column=6, value=variant.get('markupPercent', 30)).border = thin_border
            ws_models.cell(row=row, column=7, value=variant.get('price', 0)).border = thin_border
            ws_models.cell(row=row, column=8, value=variant.get('colorPreview', '')).border = thin_border
            row += 1
    
    # Auto-width columns
    for col in range(1, len(model_headers) + 1):
        ws_models.column_dimensions[chr(64 + col)].width = 15
    
    # === Sheet 2: Options ===
    ws_options = wb.create_sheet("Opcje")
    
    option_headers = ["Kategoria ID", "Kategoria (RU)", "Opcja ID", "Opcja (RU)", "Opcja (PL)", "Zakup EUR", "Marża %", "Cena PLN", "Kolor HEX"]
    for col, header in enumerate(option_headers, 1):
        cell = ws_options.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 2
    for category in prices.get('categories', []):
        cat_id = category.get('id', '')
        cat_name_ru = category.get('nameRu', category.get('name', ''))
        
        for option in category.get('options', []):
            ws_options.cell(row=row, column=1, value=cat_id).border = thin_border
            ws_options.cell(row=row, column=2, value=cat_name_ru).border = thin_border
            ws_options.cell(row=row, column=3, value=option.get('id', '')).border = thin_border
            ws_options.cell(row=row, column=4, value=option.get('nameRu', option.get('name', ''))).border = thin_border
            ws_options.cell(row=row, column=5, value=option.get('namePl', '')).border = thin_border
            ws_options.cell(row=row, column=6, value=option.get('purchasePriceEur', 0)).border = thin_border
            ws_options.cell(row=row, column=7, value=option.get('markupPercent', 30)).border = thin_border
            ws_options.cell(row=row, column=8, value=option.get('price', 0)).border = thin_border
            ws_options.cell(row=row, column=9, value=option.get('colorPreview', '')).border = thin_border
            row += 1
    
    # Auto-width columns
    for col in range(1, len(option_headers) + 1):
        ws_options.column_dimensions[chr(64 + col)].width = 18
    
    # === Sheet 3: Settings ===
    ws_settings = wb.create_sheet("Ustawienia")
    
    ws_settings.cell(row=1, column=1, value="Parametr").font = header_font
    ws_settings.cell(row=1, column=1).fill = header_fill
    ws_settings.cell(row=1, column=2, value="Wartość").font = header_font
    ws_settings.cell(row=1, column=2).fill = header_fill
    
    settings = [
        ("Waluta", prices.get('currency', 'PLN')),
        ("Symbol waluty", prices.get('currencySymbol', 'zł')),
        ("Kurs EUR", prices.get('eurRate', 4.30)),
        ("Domyślna marża %", prices.get('defaultMarkupPercent', 30)),
    ]
    
    for row, (param, value) in enumerate(settings, 2):
        ws_settings.cell(row=row, column=1, value=param).border = thin_border
        ws_settings.cell(row=row, column=2, value=value).border = thin_border
    
    ws_settings.column_dimensions['A'].width = 20
    ws_settings.column_dimensions['B'].width = 15
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"cennik_balia_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


from fastapi import UploadFile, File

@router.post("/prices/import")
async def import_prices(file: UploadFile = File(...)):
    """Import prices from Excel file"""
    from openpyxl import load_workbook
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        contents = await file.read()
        buffer = io.BytesIO(contents)
        wb = load_workbook(buffer, read_only=True)
        
        # Get current prices
        prices = await db.prices.find_one({"_id": "default"})
        if not prices:
            prices = {}
        
        updated_models = 0
        updated_options = 0
        updated_settings = 0
        
        # === Import Models ===
        if "Modele" in wb.sheetnames:
            ws = wb["Modele"]
            models = prices.get('models', [])
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                    
                model_id, name_ru, name_pl, heater_type, purchase_eur, markup, price, color = row[:8]
                
                # Find and update model
                for model in models:
                    if model.get('id') == model_id:
                        # Update variant
                        for variant in model.get('heaterVariants', []):
                            if variant.get('type') == heater_type:
                                if purchase_eur is not None:
                                    variant['purchasePriceEur'] = float(purchase_eur or 0)
                                if markup is not None:
                                    variant['markupPercent'] = float(markup or 30)
                                if price is not None:
                                    variant['price'] = float(price or 0)
                                if color:
                                    variant['colorPreview'] = str(color)
                                updated_models += 1
                        break
            
            prices['models'] = models
        
        # === Import Options ===
        if "Opcje" in wb.sheetnames:
            ws = wb["Opcje"]
            categories = prices.get('categories', [])
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                cat_id, cat_name, opt_id, opt_name_ru, opt_name_pl, purchase_eur, markup, price, color = row[:9]
                
                # Find and update option
                for category in categories:
                    if category.get('id') == cat_id:
                        for option in category.get('options', []):
                            if option.get('id') == opt_id:
                                if purchase_eur is not None:
                                    option['purchasePriceEur'] = float(purchase_eur or 0)
                                if markup is not None:
                                    option['markupPercent'] = float(markup or 30)
                                if price is not None:
                                    option['price'] = float(price or 0)
                                if color:
                                    option['colorPreview'] = str(color)
                                updated_options += 1
                                break
                        break
            
            prices['categories'] = categories
        
        # === Import Settings ===
        if "Ustawienia" in wb.sheetnames:
            ws = wb["Ustawienia"]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                    
                param, value = row[:2]
                
                if param == "Waluta":
                    prices['currency'] = str(value)
                    updated_settings += 1
                elif param == "Symbol waluty":
                    prices['currencySymbol'] = str(value)
                    updated_settings += 1
                elif param == "Kurs EUR":
                    prices['eurRate'] = float(value or 4.30)
                    updated_settings += 1
                elif param == "Domyślna marża %":
                    prices['defaultMarkupPercent'] = float(value or 30)
                    updated_settings += 1
        
        # Save to database
        await db.prices.update_one(
            {"_id": "default"},
            {"$set": prices},
            upsert=True
        )
        
        return {
            "message": "Import successful",
            "updated_models": updated_models,
            "updated_options": updated_options,
            "updated_settings": updated_settings
        }
        
    except Exception as e:
        logger.error(f"Import error: {e}")
        raise HTTPException(status_code=400, detail=f"Error importing file: {str(e)}")


@router.get("/excel-template-structure")
async def get_excel_template_structure():
    """Get Excel template structure for mapping configuration"""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'production_template.xlsx')
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Production template not found")
    
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Extract cells with labels (headers)
    cells = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.strip():
                col_letter = get_column_letter(col)
                cell_ref = f"{col_letter}{row}"
                cells.append({
                    "cell": cell_ref,
                    "value": str(cell.value).strip(),
                    "row": row,
                    "col": col
                })
    
    # Group by rows for better UI
    rows_map = {}
    for c in cells:
        row = c["row"]
        if row not in rows_map:
            rows_map[row] = []
        rows_map[row].append(c)
    
    return {
        "maxRow": ws.max_row,
        "maxCol": ws.max_column,
        "cells": cells,
        "rowsMap": rows_map
    }


@router.post("/generate-production-excel")
async def generate_production_excel(request: dict):
    """Generate production Excel with selected options marked with X using DB mapping"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    
    # Load template
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'production_template.xlsx')
    if not os.path.exists(template_path):
        raise HTTPException(status_code=500, detail="Production template not found")
    
    wb = load_workbook(template_path)
    ws = wb.active
    
    # Style for X marks
    x_font = Font(bold=True, size=14)
    x_align = Alignment(horizontal='center', vertical='center')
    
    # Get prices data with excelCell mappings from DB
    prices = await db.prices.find_one({"_id": "default"})
    if not prices:
        raise HTTPException(status_code=500, detail="Prices not found")
    
    # Build mapping from DB
    # Models mapping: model_id -> excelCell
    model_mapping = {}
    heater_mapping = {}  # heaterVariant id -> excelCell
    for model in prices.get('models', []):
        model_id = model.get('id')
        if model_id and model.get('excelCell'):
            model_mapping[model_id] = model.get('excelCell')
        # HeaterVariants mapping
        for hv in model.get('heaterVariants', []):
            hv_id = hv.get('id')
            if hv_id and hv.get('excelCell'):
                heater_mapping[hv_id] = hv.get('excelCell')
    
    # Options mapping: option_id -> excelCell
    option_mapping = {}
    for cat in prices.get('categories', []):
        for opt in cat.get('options', []):
            opt_id = opt.get('id')
            if opt_id and opt.get('excelCell'):
                option_mapping[opt_id] = opt.get('excelCell')
    
    logger.info(f"Loaded mappings - Models: {len(model_mapping)}, Heaters: {len(heater_mapping)}, Options: {len(option_mapping)}")
    
    def mark_cell(cell_ref):
        """Put X in specified cell"""
        if cell_ref:
            ws[cell_ref] = 'X'
            ws[cell_ref].font = x_font
            ws[cell_ref].alignment = x_align
            logger.info(f"Marked cell: {cell_ref}")
    
    # Customer data
    ws['B2'] = request.get('fullName', '')
    ws['B4'] = request.get('fullAddress', '')
    
    # Process model
    model_id = request.get('modelId', '')
    if model_id and model_id in model_mapping:
        mark_cell(model_mapping[model_id])
        logger.info(f"Model {model_id} -> {model_mapping[model_id]}")
    
    # Process heater variant
    heater_variant_id = request.get('selectedHeaterVariantId', '') or request.get('heaterVariantId', '')
    heater_type = request.get('heaterType', '')  # "integrated" or "external"
    
    # First try by ID
    if heater_variant_id and heater_variant_id in heater_mapping:
        mark_cell(heater_mapping[heater_variant_id])
        logger.info(f"Heater by ID {heater_variant_id} -> {heater_mapping[heater_variant_id]}")
    # Then try to find by model_id + heater_type combination
    elif model_id and heater_type:
        # Try common patterns for heater variant ID
        possible_ids = [
            f"{model_id}_{heater_type}",  # round_200_integrated
            f"{model_id}_{heater_type[:3]}",  # round_200_int
            heater_type,  # just "integrated" or "external"
        ]
        for pid in possible_ids:
            if pid in heater_mapping:
                mark_cell(heater_mapping[pid])
                logger.info(f"Heater by pattern {pid} -> {heater_mapping[pid]}")
                break
        else:
            logger.warning(f"No heater mapping found for model={model_id}, type={heater_type}, id={heater_variant_id}")
    
    # Process selected options from selections dict
    selections = request.get('selections', {})
    for cat_id, selection in selections.items():
        if isinstance(selection, dict):
            # Checkbox type - multiple selections
            for opt_id, is_selected in selection.items():
                if is_selected and opt_id in option_mapping:
                    mark_cell(option_mapping[opt_id])
        elif isinstance(selection, str):
            # Radio/dropdown type - single selection
            if selection in option_mapping:
                mark_cell(option_mapping[selection])
    
    # Also check selectedOptions array for backward compatibility
    selected_options = request.get('selectedOptions', [])
    for opt in selected_options:
        opt_id = opt.get('id', '') if isinstance(opt, dict) else str(opt)
        if opt_id in option_mapping:
            mark_cell(option_mapping[opt_id])
    
    # Add notes if present
    notes = request.get('notes', '')
    if notes:
        ws['B18'] = notes
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename
    safe_name = (request.get('fullName', 'Zamowienie') or 'Zamowienie').replace(' ', '_')
    safe_name = ''.join(c for c in safe_name if c.isalnum() or c == '_')
    order_id = request.get('orderId', request.get('id', 'new'))
    filename = f"TechSpec_{safe_name}_{order_id}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={quote(filename)}"}
    )


@router.post("/clear-images")
async def clear_all_images():
    """Clear all image URLs from models and options"""
    prices = await db.prices.find_one({"_id": "default"})
    if not prices:
        return {"message": "No prices found"}
    
    # Clear model images
    models = prices.get('models', [])
    for model in models:
        model['imageUrl'] = ''
    
    # Clear category and option images
    categories = prices.get('categories', [])
    for category in categories:
        category['imageUrl'] = ''
        for option in category.get('options', []):
            option['imageUrl'] = ''
    
    # Update database
    await db.prices.update_one(
        {"_id": "default"},
        {"$set": {"models": models, "categories": categories}}
    )
    
    return {"message": "All Balia images cleared successfully", "models_cleared": len(models), "categories_cleared": len(categories)}


@router.post("/orders", response_model=Order)
async def create_order(order: Order):
    """Create a new order"""
    order_dict = order.model_dump()
    await db.orders.insert_one(order_dict)
    
    # Send Telegram notification with PDF for new balia order
    try:
        # Generate PDF for the order
        from models.balia import PDFRequest
        pdf_request = PDFRequest(
            orderId=order_dict.get('id'),
            fullName=order_dict.get('fullName', ''),
            phoneNumber=order_dict.get('phoneNumber', ''),
            fullAddress=order_dict.get('fullAddress', order_dict.get('address', '')),
            orderDate=order_dict.get('orderDate', order_dict.get('createdAt', datetime.now(timezone.utc).isoformat())),
            email=order_dict.get('email'),
            notes=order_dict.get('notes'),
            modelName=order_dict.get('modelName'),
            heaterType=order_dict.get('heaterType'),
            modelSpecs=order_dict.get('modelSpecs'),
            selectedOptions=order_dict.get('selectedOptions', []),
            total=order_dict.get('total'),
            currency=order_dict.get('currency', 'PLN'),
            currencySymbol=order_dict.get('currencySymbol', 'zł')
        )
        pdf_data = await generate_pdf_bytes(pdf_request)
        await notify_new_order(order_dict, order_type='balia', is_web_order=False, pdf_data=pdf_data)
    except Exception as e:
        logger.warning(f"Failed to send Telegram notification with PDF for order: {e}")
        # Fallback: try to send without PDF
        try:
            await notify_new_order(order_dict, order_type='balia', is_web_order=False)
        except:
            pass
    
    return order


@router.get("/orders", response_model=List[Order])
async def get_orders():
    """Get all orders"""
    orders = await db.orders.find({}, {"_id": 0}).to_list(1000)
    return orders


@router.get("/orders/{order_id}")
async def get_order(order_id: str):
    """Get a single order by ID"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.put("/orders/{order_id}")
async def update_order(order_id: str, order: Order):
    """Update an existing order"""
    order_dict = order.model_dump()
    result = await db.orders.update_one(
        {"id": order_id},
        {"$set": order_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.delete("/orders/{order_id}")
async def delete_order(order_id: str):
    """Delete an order"""
    result = await db.orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"message": "Order deleted successfully"}


async def generate_pdf_bytes(request: PDFRequest) -> bytes:
    """Generate PDF and return as bytes (for Telegram sending)"""
    import urllib.request
    import base64
    from PIL import Image as PILImage
    
    async def load_image_from_mongodb(image_url: str) -> bytes:
        """Load image from MongoDB by extracting ID from URL"""
        if not image_url or '/api/uploads/' not in image_url:
            return None
        try:
            filename = image_url.split('/api/uploads/')[-1]
            file_id = filename.rsplit('.', 1)[0] if '.' in filename else filename
            image_doc = await db.images.find_one({"id": file_id})
            if image_doc:
                return base64.b64decode(image_doc["content"])
        except Exception as e:
            logger.warning(f"Could not load image from MongoDB: {e}")
        return None
    
    buffer = io.BytesIO()
    
    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
    except Exception as e:
        logger.warning(f"Could not register fonts: {e}")
    
    # Colors - Blue theme for Balia (water theme)
    BLUE = colors.HexColor('#2563EB')
    BLUE_LIGHT = colors.HexColor('#EFF6FF')
    BLUE_DARK = colors.HexColor('#1E40AF')
    BLUE_BORDER = colors.HexColor('#93C5FD')
    GREEN = colors.HexColor('#059669')
    GREEN_LIGHT = colors.HexColor('#ECFDF5')
    TEXT_COLOR = colors.HexColor('#1F2937')
    MUTED = colors.HexColor('#6B7280')
    WHITE = colors.white
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=10*mm,
        bottomMargin=15*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Generate offer number
    offer_number = request.orderId if request.orderId else f"WM-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Custom styles
    title_style = ParagraphStyle('Title', fontName='DejaVuSans-Bold', fontSize=22, textColor=BLUE_DARK, alignment=TA_CENTER)
    subtitle_style = ParagraphStyle('Subtitle', fontName='DejaVuSans', fontSize=10, textColor=MUTED, alignment=TA_CENTER)
    section_style = ParagraphStyle('Section', fontName='DejaVuSans-Bold', fontSize=13, textColor=BLUE_DARK, spaceBefore=12)
    normal_style = ParagraphStyle('Normal', fontName='DejaVuSans', fontSize=10, textColor=TEXT_COLOR)
    
    # Header
    elements.append(Paragraph("OFERTA CENOWA", title_style))
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(f"Nr oferty: {offer_number} | Data: {datetime.now().strftime('%d.%m.%Y')}", subtitle_style))
    elements.append(Spacer(1, 6*mm))
    
    # Customer info section
    elements.append(Paragraph("📋 DANE KLIENTA", section_style))
    customer_data = [
        ["Imię i nazwisko:", request.fullName or "-"],
        ["Telefon:", request.phoneNumber or "-"],
        ["Email:", getattr(request, 'email', None) or "-"],
    ]
    address = getattr(request, 'fullAddress', None) or getattr(request, 'address', None)
    if address:
        customer_data.append(["Adres:", address])
    notes = getattr(request, 'notes', None)
    if notes:
        customer_data.append(["Uwagi:", notes])
    
    customer_table = Table(customer_data, colWidths=[50*mm, 125*mm])
    customer_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'DejaVuSans'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), MUTED),
        ('TEXTCOLOR', (1, 0), (1, -1), TEXT_COLOR),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(customer_table)
    elements.append(Spacer(1, 6*mm))
    
    # Model section
    elements.append(Paragraph("🛁 WYBRANY MODEL", section_style))
    
    heater_type = request.heaterType or 'external'
    heater_type_name = "Zintegrowany" if heater_type == 'integrated' else "Zewnętrzny"
    
    model_info = [
        ["Model:", request.modelName or "-"],
        ["Piec:", heater_type_name],
    ]
    
    # Add specs if available
    if request.modelSpecs:
        specs = request.modelSpecs
        if specs.get('dimensions'):
            model_info.append(["Wymiary:", specs['dimensions']])
        if specs.get('innerDiameter'):
            model_info.append(["Średnica wewnętrzna:", f"{specs['innerDiameter']} cm"])
        if specs.get('depth'):
            model_info.append(["Głębokość:", f"{specs['depth']} cm"])
        if specs.get('capacity'):
            model_info.append(["Pojemność:", f"{specs['capacity']} L"])
        if specs.get('persons'):
            model_info.append(["Liczba osób:", f"{specs['persons']}"])
    
    model_table = Table(model_info, colWidths=[50*mm, 125*mm])
    model_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'DejaVuSans'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), MUTED),
        ('TEXTCOLOR', (1, 0), (1, -1), TEXT_COLOR),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(model_table)
    elements.append(Spacer(1, 6*mm))
    
    # Options section
    if request.selectedOptions:
        elements.append(Paragraph("📦 WYBRANE OPCJE", section_style))
        
        options_data = [["Opcja", "Cena"]]
        for opt in request.selectedOptions:
            opt_name = opt.get('optionName') or opt.get('name') or opt.get('namePl', '-')
            opt_price = opt.get('optionPrice') or opt.get('price', 0)
            currency_symbol = request.currencySymbol or 'zł'
            options_data.append([opt_name, f"{opt_price:,.0f} {currency_symbol}".replace(",", " ")])
        
        options_table = Table(options_data, colWidths=[125*mm, 50*mm])
        options_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, 0), BLUE_LIGHT),
            ('TEXTCOLOR', (0, 0), (-1, 0), BLUE_DARK),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, BLUE_BORDER),
        ]))
        elements.append(options_table)
        elements.append(Spacer(1, 6*mm))
    
    # Total section
    elements.append(Paragraph("💰 PODSUMOWANIE", section_style))
    
    currency_symbol = request.currencySymbol or 'zł'
    total = request.total or 0
    total_formatted = f"{total:,.0f}".replace(",", " ")
    
    total_data = [
        ["RAZEM DO ZAPŁATY:", f"{total_formatted} {currency_symbol}"]
    ]
    
    total_table = Table(total_data, colWidths=[125*mm, 50*mm])
    total_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('BACKGROUND', (0, 0), (-1, -1), GREEN_LIGHT),
        ('TEXTCOLOR', (0, 0), (0, -1), TEXT_COLOR),
        ('TEXTCOLOR', (1, 0), (1, -1), GREEN),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOX', (0, 0), (-1, -1), 1, GREEN),
    ]))
    elements.append(total_table)
    elements.append(Spacer(1, 8*mm))
    
    # Footer
    footer_style = ParagraphStyle('Footer', fontName='DejaVuSans', fontSize=8, textColor=MUTED, alignment=TA_CENTER)
    elements.append(Paragraph("Oferta ważna 14 dni od daty wystawienia.", footer_style))
    elements.append(Paragraph("Oferta nie stanowi oferty handlowej w rozumieniu Kodeksu Cywilnego.", footer_style))
    
    doc.build(elements)
    
    pdf_data = buffer.getvalue()
    buffer.close()
    
    return pdf_data


@router.post("/generate-pdf")
async def generate_pdf(request: PDFRequest):
    """Generate professional PDF order form for Balia - Polish only"""
    import urllib.request
    import base64
    from PIL import Image as PILImage
    
    async def load_image_from_mongodb(image_url: str) -> bytes:
        """Load image from MongoDB by extracting ID from URL"""
        if not image_url or '/api/uploads/' not in image_url:
            return None
        try:
            filename = image_url.split('/api/uploads/')[-1]
            file_id = filename.rsplit('.', 1)[0] if '.' in filename else filename
            image_doc = await db.images.find_one({"id": file_id})
            if image_doc:
                return base64.b64decode(image_doc["content"])
        except Exception as e:
            logger.warning(f"Could not load image from MongoDB: {e}")
        return None
    
    buffer = io.BytesIO()
    
    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
    except Exception as e:
        logger.warning(f"Could not register fonts: {e}")
    
    # Colors - Blue theme for Balia (water theme)
    BLUE = colors.HexColor('#2563EB')
    BLUE_LIGHT = colors.HexColor('#EFF6FF')
    BLUE_DARK = colors.HexColor('#1E40AF')
    BLUE_BORDER = colors.HexColor('#93C5FD')
    GREEN = colors.HexColor('#059669')
    GREEN_LIGHT = colors.HexColor('#ECFDF5')
    TEXT_COLOR = colors.HexColor('#1F2937')
    MUTED = colors.HexColor('#6B7280')
    WHITE = colors.white
    
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                          rightMargin=20, leftMargin=20,
                          topMargin=20, bottomMargin=20)
    
    elements = []
    styles = getSampleStyleSheet()
    
    section_title_style = ParagraphStyle(
        'SectionTitle',
        fontName='DejaVuSans-Bold',
        fontSize=13,
        textColor=BLUE_DARK,
    )
    
    normal_style = ParagraphStyle(
        'Normal',
        fontName='DejaVuSans',
        fontSize=9,
        textColor=TEXT_COLOR,
    )
    
    # Calculate dates
    current_date = datetime.now().strftime('%d.%m.%Y')
    valid_until = (datetime.now() + timedelta(days=30)).strftime('%d.%m.%Y')
    
    # Use orderId if provided, otherwise generate new number
    if request.orderId and request.orderId.startswith('WMB-'):
        offer_number = request.orderId
    else:
        offer_number = f"WMB-{datetime.now().strftime('%d-%m-%Y-%H%M%S')}"
    currency = request.currency or 'EUR'
    
    # Load model image - first try from request, then fetch from DB based on heaterType
    model_img = None
    model_image_url = getattr(request, 'modelImageUrl', None)
    
    # If modelImageUrl is empty or not provided, get it from DB based on modelId and heaterType
    if not model_image_url and request.modelId:
        try:
            prices_data = await db.prices.find_one({"_id": "default"})
            if prices_data:
                for m in prices_data.get('models', []):
                    if m.get('id') == request.modelId:
                        heater_type = getattr(request, 'heaterType', None) or 'integrated'
                        # Look for image in heaterVariants based on selected heater type
                        heater_variants = m.get('heaterVariants', [])
                        if heater_variants:
                            # Find the variant matching the selected heater type
                            for variant in heater_variants:
                                if variant.get('type') == heater_type and variant.get('imageUrl'):
                                    model_image_url = variant.get('imageUrl')
                                    logger.info(f"Found model image from heaterVariant '{heater_type}': {model_image_url[:50]}...")
                                    break
                            # Fallback to first variant with image
                            if not model_image_url:
                                for variant in heater_variants:
                                    if variant.get('imageUrl'):
                                        model_image_url = variant.get('imageUrl')
                                        logger.info(f"Using fallback heaterVariant image: {model_image_url[:50]}...")
                                        break
                        # Ultimate fallback to model's own imageUrl
                        if not model_image_url:
                            model_image_url = m.get('imageUrl', '')
                            if model_image_url:
                                logger.info(f"Using model's own imageUrl: {model_image_url[:50]}...")
                        break
        except Exception as e:
            logger.warning(f"Error fetching model image from DB: {e}")
    
    if model_image_url:
        try:
            img_data = None
            
            # Handle Base64 data URI
            if model_image_url.startswith('data:image'):
                try:
                    # Extract base64 part from data URI
                    base64_data = model_image_url.split(',', 1)[1]
                    img_data = base64.b64decode(base64_data)
                    logger.info(f"Decoded Base64 model image, size: {len(img_data)} bytes")
                except Exception as e:
                    logger.warning(f"Could not decode Base64 model image: {e}")
            
            # Try loading from MongoDB
            if not img_data and '/api/uploads/' in model_image_url:
                img_data = await load_image_from_mongodb(model_image_url)
                if img_data:
                    logger.info(f"Loaded model image from MongoDB")
            
            # Fallback to HTTP download for external URLs
            if not img_data and model_image_url.startswith('http'):
                try:
                    img_data = urllib.request.urlopen(model_image_url, timeout=5).read()
                    logger.info(f"Downloaded model image from URL: {model_image_url[:80]}...")
                except Exception as e:
                    logger.warning(f"Could not download image from URL: {e}")
            
            if img_data:
                # Get original image dimensions to preserve aspect ratio
                img_buffer = io.BytesIO(img_data)
                pil_img = PILImage.open(img_buffer)
                orig_width, orig_height = pil_img.size
                
                # Calculate scaled dimensions (max width 160, preserve ratio)
                max_width = 160
                max_height = 120
                ratio = min(max_width / orig_width, max_height / orig_height)
                new_width = orig_width * ratio
                new_height = orig_height * ratio
                
                img_buffer.seek(0)
                model_img = RLImage(img_buffer, width=new_width, height=new_height)
        except Exception as e:
            logger.warning(f"Could not load model image: {e}")
    
    # ========== HEADER - styled WM-BALIA text ==========
    logo_style = ParagraphStyle(
        'LogoStyle',
        fontName='DejaVuSans-Bold',
        fontSize=28,
        textColor=BLUE_DARK,
        leading=32,
    )
    logo_cell = Paragraph('<font color="#2563EB">WM</font><font color="#1E40AF">-BALIA</font>', logo_style)
    
    header_data = [[
        logo_cell,
        '',
        Paragraph('''<b>OFERTA HANDLOWA</b><br/>
        <font size="9" color="#6B7280">Tel: +48 732 111 111</font><br/>
        <font size="9" color="#6B7280">Email: wmbalia@gmail.com</font><br/>
        <font size="9" color="#6B7280">www.wm-balia.pl</font>''',
        ParagraphStyle('HeaderRight', fontName='DejaVuSans', fontSize=16, alignment=TA_RIGHT, textColor=BLUE))
    ]]
    header_table = Table(header_data, colWidths=[200, 130, 200])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BLUE_LIGHT),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (0, 0), 10),
    ]))
    elements.append(header_table)
    
    # Divider line
    elements.append(Spacer(1, 4))
    elements.append(Table([['']], colWidths=[530], rowHeights=[3], style=[('BACKGROUND', (0,0), (0,0), BLUE)]))
    elements.append(Spacer(1, 10))
    
    # ========== CLIENT + OFFER INFO ==========
    client_info = Paragraph(f'''<b>DANE KLIENTA:</b><br/>
    Imię i nazwisko: {request.fullName}<br/>
    Telefon: {request.phoneNumber}<br/>
    Adres: {request.fullAddress}''', 
    ParagraphStyle('ClientInfo', fontName='DejaVuSans', fontSize=9, textColor=TEXT_COLOR))
    
    offer_info = Paragraph(f'''<b>INFORMACJE O OFERCIE:</b><br/>
    Data wystawienia: {current_date}<br/>
    Ważność oferty: {valid_until}<br/>
    <b>Nr oferty: {offer_number}</b>''',
    ParagraphStyle('OfferInfo', fontName='DejaVuSans', fontSize=9, textColor=TEXT_COLOR, alignment=TA_RIGHT))
    
    info_table = Table([[client_info, offer_info]], colWidths=[265, 265])
    info_table.setStyle(TableStyle([
        ('BOX', (0, 0), (-1, -1), 1, BLUE),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 12))
    
    # ========== MODEL SECTION WITH IMAGE - Polish name ==========
    if request.modelName or request.modelId:
        # Get Polish model name and specs from DB
        model_name = request.modelName
        model_price = request.modelPrice or 0
        model_specs = getattr(request, 'modelSpecs', None) or {}
        
        # Try to get Polish name and specs from database if not provided
        if request.modelId:
            prices_data_for_model = await db.prices.find_one({"_id": "default"})
            if prices_data_for_model:
                for m in prices_data_for_model.get('models', []):
                    if m.get('id') == request.modelId:
                        model_name = m.get('namePl') or m.get('name', model_name)
                        # Use request specs if provided, otherwise fallback to DB
                        if not model_specs:
                            model_specs = m.get('specs', {})
                        break
        
        # Create model info content with specs
        specs_text = ""
        if model_specs:
            specs_lines = []
            if model_specs.get('outerDiameter'):
                specs_lines.append(f"Średnica zewnętrzna: {model_specs['outerDiameter']}")
            if model_specs.get('innerDiameter'):
                specs_lines.append(f"Średnica wewnętrzna: {model_specs['innerDiameter']}")
            if model_specs.get('dimensions'):
                specs_lines.append(f"Wymiary: {model_specs['dimensions']}")
            if model_specs.get('depth'):
                specs_lines.append(f"Głębokość: {model_specs['depth']}")
            if model_specs.get('volume'):
                specs_lines.append(f"Pojemność: {model_specs['volume']}")
            if model_specs.get('seats') and model_specs.get('seats') > 0:
                specs_lines.append(f"Ilość miejsc: {model_specs['seats']}")
            if model_specs.get('totalHeight') and model_specs.get('totalHeight') not in [0, '0']:
                specs_lines.append(f"Wysokość całkowita: {model_specs['totalHeight']}")
            if model_specs.get('heaterPower') and model_specs.get('heaterPower') not in [0, '0']:
                specs_lines.append(f"Moc pieca: {model_specs['heaterPower']}")
            if model_specs.get('weight'):
                specs_lines.append(f"Waga (pusta): {model_specs['weight']}")
            
            if specs_lines:
                specs_text = "<br/><br/><font size='9' color='#6B7280'>" + " | ".join(specs_lines) + "</font>"
        
        # Get heater type info - try multiple ways to access the field
        heater_type = None
        heater_type_name = None
        
        # Try accessing as Pydantic model attribute
        if hasattr(request, 'heaterType') and request.heaterType:
            heater_type = request.heaterType
        # Try accessing from model_dump()
        request_dict = request.model_dump() if hasattr(request, 'model_dump') else {}
        if not heater_type and request_dict.get('heaterType'):
            heater_type = request_dict.get('heaterType')
        
        if hasattr(request, 'heaterTypeName') and request.heaterTypeName:
            heater_type_name = request.heaterTypeName
        elif request_dict.get('heaterTypeName'):
            heater_type_name = request_dict.get('heaterTypeName')
        
        # Generate heater type name from type if not provided
        if not heater_type_name and heater_type:
            heater_type_name = 'Piec zintegrowany' if heater_type == 'integrated' else 'Piec zewnętrzny'
        
        logger.info(f"PDF heater info: type={heater_type}, name={heater_type_name}")
        
        heater_text = f"<br/><font size='10' color='#059669'><b>Typ pieca: {heater_type_name}</b></font>" if heater_type_name else ""
        
        model_text = Paragraph(f'''<b><font size="14" color="#1E40AF">WYBRANY MODEL</font></b><br/><br/>
        <font size="12"><b>{model_name}</b></font>{heater_text}<br/><br/>
        <font size="11">Cena bazowa: <b>{model_price:,.0f} {currency}</b></font>{specs_text}'''.replace(',', ' '),
        ParagraphStyle('ModelText', fontName='DejaVuSans', fontSize=11, textColor=TEXT_COLOR))
        
        if model_img:
            model_data = [[model_img, model_text]]
            model_table = Table(model_data, colWidths=[170, 350])
        else:
            model_data = [[model_text]]
            model_table = Table(model_data, colWidths=[520])
        
        model_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BLUE_LIGHT),
            ('BOX', (0, 0), (-1, -1), 2, BLUE),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ]))
        elements.append(model_table)
        elements.append(Spacer(1, 12))
    
    # ========== SELECTED OPTIONS WITH IMAGES - Always in Polish ==========
    # Get admin gifts list if available
    admin_gifts = getattr(request, 'adminGifts', []) or []
    
    if request.selectedOptions and len(request.selectedOptions) > 0:
        elements.append(Paragraph('<b>WYBRANE OPCJE</b>', section_title_style))
        elements.append(Spacer(1, 6))
        
        # Load prices from DB to get Polish names and images
        prices_data = await db.prices.find_one({"_id": "default"})
        categories_map = {}
        if prices_data:
            for cat in prices_data.get('categories', []):
                cat_id = cat.get('id')
                cat_name_pl = cat.get('namePl') or cat.get('name', '')
                cat_image_url = cat.get('imageUrl', '')
                options_map = {}
                for opt in cat.get('options', []):
                    opt_id = opt.get('id')
                    opt_name_pl = opt.get('namePl') or opt.get('name', '')
                    opt_image_url = opt.get('imageUrl', '')
                    options_map[opt_id] = {
                        'name': opt_name_pl,
                        'imageUrl': opt_image_url
                    }
                categories_map[cat_id] = {
                    'name': cat_name_pl, 
                    'imageUrl': cat_image_url,
                    'options': options_map
                }
        
        # Helper function to load and resize option image
        async def load_option_image(image_url: str, max_width: int = 60, max_height: int = 45):
            """Load option image from MongoDB, base64, or external URL"""
            if not image_url:
                return None
            try:
                img_data = None
                
                # Handle base64 images
                if image_url.startswith('data:image'):
                    try:
                        # Extract base64 data after the comma
                        base64_data = image_url.split(',', 1)[1] if ',' in image_url else image_url
                        import base64
                        img_data = base64.b64decode(base64_data)
                        logger.info(f"Decoded base64 image: {len(img_data)} bytes")
                    except Exception as e:
                        logger.warning(f"Failed to decode base64 image: {e}")
                
                # Try loading from MongoDB
                if not img_data and '/api/uploads/' in image_url:
                    img_data = await load_image_from_mongodb(image_url)
                    if img_data:
                        logger.info(f"Loaded option image from MongoDB: {len(img_data)} bytes")
                
                # If not in MongoDB and it's a relative URL, try full external URL
                if not img_data and '/api/uploads/' in image_url:
                    try:
                        # Convert relative URL to absolute using API_URL env var
                        import os
                        base_url = os.environ.get('API_BASE_URL', os.environ.get('REACT_APP_BACKEND_URL', ''))
                        full_url = f"{base_url}{image_url}"
                        img_data = urllib.request.urlopen(full_url, timeout=10).read()
                        # Check if it's actually image data (not JSON error)
                        if img_data and len(img_data) > 100 and not img_data.startswith(b'{'):
                            logger.info(f"Downloaded option image from external URL: {len(img_data)} bytes")
                        else:
                            logger.warning(f"External URL returned non-image data: {image_url}")
                            img_data = None
                    except Exception as e:
                        logger.warning(f"Failed to download from external URL: {e}")
                
                # Try direct HTTP download for external URLs
                if not img_data and image_url.startswith('http'):
                    try:
                        img_data = urllib.request.urlopen(image_url, timeout=10).read()
                        if img_data and len(img_data) > 100:
                            logger.info(f"Downloaded option image from URL: {len(img_data)} bytes")
                    except Exception as e:
                        logger.warning(f"Failed to download option image: {e}")
                
                if img_data:
                    # Compress and resize image
                    img_buffer = io.BytesIO(img_data)
                    pil_img = PILImage.open(img_buffer)
                    
                    # Convert to RGB if necessary
                    if pil_img.mode in ('RGBA', 'P'):
                        pil_img = pil_img.convert('RGB')
                    
                    # Resize preserving aspect ratio
                    orig_width, orig_height = pil_img.size
                    ratio = min(max_width / orig_width, max_height / orig_height)
                    new_width = int(orig_width * ratio)
                    new_height = int(orig_height * ratio)
                    pil_img = pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
                    
                    # Compress
                    output = io.BytesIO()
                    pil_img.save(output, format='JPEG', quality=70, optimize=True)
                    output.seek(0)
                    
                    logger.info(f"Option image processed: {new_width}x{new_height}")
                    return RLImage(output, width=new_width, height=new_height)
            except Exception as e:
                logger.warning(f"Could not load option image: {e}")
            return None
        
        options_data = [['', 'Kategoria', 'Wybrana opcja', 'Cena']]
        total_options_price = 0
        gifts_total = 0
        gift_rows = []  # Track which rows are gifts for styling
        
        # Style for gift price with strikethrough
        gift_price_style = ParagraphStyle(
            'GiftPrice',
            fontName='DejaVuSans',
            fontSize=9,
            textColor=colors.HexColor('#059669'),
            alignment=TA_RIGHT
        )
        normal_price_style = ParagraphStyle(
            'NormalPrice',
            fontName='DejaVuSans',
            fontSize=9,
            textColor=TEXT_COLOR,
            alignment=TA_RIGHT
        )
        
        for idx, opt in enumerate(request.selectedOptions):
            cat_id = opt.get('categoryId', '')
            opt_id = opt.get('optionId', '') or opt.get('id', '')
            price = opt.get('price', 0)
            is_not_selected = opt.get('notSelected', False) or opt_id is None or 'not_selected' in str(opt.get('id', ''))
            
            # Check if this option is a gift
            is_gift = opt_id in admin_gifts if opt_id else False
            
            if is_gift:
                gifts_total += price
            elif not is_not_selected:
                total_options_price += price
            
            # Get Polish names and images from DB, fallback to provided names
            cat_info = categories_map.get(cat_id, {})
            cat_name = cat_info.get('name', opt.get('categoryName', ''))
            
            if is_not_selected:
                # Not selected - show "Bez X" in gray (use name from frontend)
                opt_name = opt.get('optionName', '') or opt.get('name', '') or 'Nie wybrano'
                img_cell = ''
                price_text = '-'
                price_cell = Paragraph(price_text, ParagraphStyle(
                    'NotSelected',
                    fontName='DejaVuSans',
                    fontSize=9,
                    textColor=colors.HexColor('#999999'),
                    alignment=TA_RIGHT
                ))
                not_selected_style = ParagraphStyle(
                    'NotSelectedText',
                    fontName='DejaVuSans',
                    fontSize=9,
                    textColor=colors.HexColor('#999999')
                )
                options_data.append([img_cell, cat_name, Paragraph(opt_name, not_selected_style), price_cell])
            else:
                opt_info = cat_info.get('options', {}).get(opt_id, {})
                opt_name = opt_info.get('name', opt.get('optionName', '') or opt.get('name', ''))
                
                # Get image URL - option image first, then category image as fallback
                opt_image_url = opt_info.get('imageUrl', '')
                cat_image_url = cat_info.get('imageUrl', '')
                image_url = opt_image_url or cat_image_url
                
                logger.info(f"Option {opt_id}: opt_image_url={opt_image_url}, cat_image_url={cat_image_url}, using={image_url}")
                
                # Load image
                option_img = await load_option_image(image_url)
                img_cell = option_img if option_img else ''
                
                if option_img:
                    logger.info(f"Successfully loaded image for option {opt_id}")
                
                if is_gift:
                    # Show as gift with strikethrough price and WM-Group label
                    price_text = f"<strike>{price:,.0f} {currency}</strike><br/>🎁 Prezent od WM-Group".replace(',', ' ')
                    price_cell = Paragraph(price_text, gift_price_style)
                    gift_rows.append(idx + 1)  # +1 because of header row
                else:
                    price_text = f"+{price:,.0f} {currency}".replace(',', ' ') if price > 0 else 'W cenie'
                    price_cell = Paragraph(price_text, normal_price_style)
                
                options_data.append([img_cell, cat_name, opt_name, price_cell])
        
        # Add subtotal row for options
        if total_options_price > 0 or gifts_total > 0:
            options_data.append(['', '', 'Opcje razem:', f"+{total_options_price:,.0f} {currency}".replace(',', ' ')])
        
        options_table = Table(options_data, colWidths=[70, 130, 230, 90])
        
        # Build table style
        table_style = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            # Body
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_COLOR),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Price column
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Image column
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),  # Vertical align all cells
            # Subtotal row
            ('FONTNAME', (2, -1), (-1, -1), 'DejaVuSans-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), BLUE_LIGHT),
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, BLUE_BORDER),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            # Alternate row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [WHITE, colors.HexColor('#F9FAFB')]),
        ]
        
        # Highlight gift rows with green background
        for gift_row in gift_rows:
            table_style.append(('BACKGROUND', (0, gift_row), (-1, gift_row), colors.HexColor('#ECFDF5')))
            table_style.append(('TEXTCOLOR', (1, gift_row), (2, gift_row), colors.HexColor('#059669')))
        
        options_table.setStyle(TableStyle(table_style))
        elements.append(options_table)
        elements.append(Spacer(1, 12))
    
    # ========== NOTES ==========
    if request.notes:
        elements.append(Paragraph('<b>UWAGI DODATKOWE</b>', section_title_style))
        elements.append(Spacer(1, 4))
        notes_para = Paragraph(request.notes, normal_style)
        notes_table = Table([[notes_para]], colWidths=[520])
        notes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
            ('BOX', (0, 0), (-1, -1), 1, MUTED),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(notes_table)
        elements.append(Spacer(1, 12))
    
    # ========== DISCOUNT SECTION (if applicable) ==========
    discount_percent = getattr(request, 'discountPercent', 0) or 0
    subtotal = getattr(request, 'subtotal', request.total / (1 - discount_percent/100) if discount_percent else request.total) or request.total
    total_after_discount = request.total
    
    if discount_percent > 0:
        savings = subtotal - total_after_discount
        discount_section = Paragraph(f'''<b><font color="#059669" size="13">ZASTOSOWANY RABAT</font></b><br/><br/>
        <font size="11">Cena przed rabatem: <b>{subtotal:,.2f} {currency}</b></font><br/>
        <font size="12" color="#059669"><b>Rabat: {discount_percent:.0f}%</b></font><br/>
        <font size="11">Kwota rabatu: <b>-{savings:,.2f} {currency}</b></font><br/><br/>
        <font size="10" color="#059669"><i>Oszczędzasz: {savings:,.2f} {currency}</i></font>'''.replace(',', ' '),
        ParagraphStyle('DiscountSection', fontName='DejaVuSans', fontSize=11))
        
        discount_table = Table([[discount_section]], colWidths=[520])
        discount_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), GREEN_LIGHT),
            ('BOX', (0, 0), (-1, -1), 2, GREEN),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ]))
        elements.append(discount_table)
        elements.append(Spacer(1, 12))
    
    # ========== TOTAL ==========
    elements.append(Spacer(1, 10))
    
    # Build total section with discount info if applicable
    if discount_percent > 0:
        total_html = f'''<font size="14"><b>SUMA DO ZAPŁATY:</b></font><br/>
        <font size="9" color="#EFF6FF">Rabat: {discount_percent:.0f}% (cena przed rabatem: {subtotal:,.2f} {currency})</font>'''.replace(',', ' ')
    else:
        total_html = f'''<font size="14"><b>SUMA DO ZAPŁATY:</b></font>'''
    
    total_data = [[
        Paragraph(total_html, 
                 ParagraphStyle('TotalLabel', fontName='DejaVuSans-Bold', fontSize=14, textColor=WHITE)),
        Paragraph(f'''<font size="18"><b>{total_after_discount:,.2f} {currency}</b></font>'''.replace(',', ' '),
                 ParagraphStyle('TotalValue', fontName='DejaVuSans-Bold', fontSize=18, textColor=WHITE, alignment=TA_RIGHT))
    ]]
    total_table = Table(total_data, colWidths=[300, 220])
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), BLUE),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 14),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 14),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
    ]))
    elements.append(total_table)
    
    # ========== FOOTER ==========
    elements.append(Spacer(1, 20))
    footer_text = Paragraph('''<font size="8" color="#6B7280">
    Dziękujemy za zainteresowanie naszą ofertą. W razie pytań prosimy o kontakt.<br/>
    Oferta nie stanowi oferty handlowej w rozumieniu Kodeksu Cywilnego.
    </font>''', ParagraphStyle('Footer', fontName='DejaVuSans', fontSize=8, textColor=MUTED, alignment=TA_CENTER))
    elements.append(footer_text)
    
    doc.build(elements)
    
    pdf_data = buffer.getvalue()
    buffer.close()
    
    # Generate filename: BALIA_ClientName_OrderId
    try:
        # Keep name with cyrillic/polish chars, just replace spaces and remove unsafe filename chars
        safe_name = request.fullName.replace(' ', '_') if request.fullName else 'Klient'
        # Remove characters that are unsafe in filenames
        safe_name = ''.join(c for c in safe_name if c not in '<>:"/\\|?*')
        if not safe_name:
            safe_name = "Klient"
    except:
        safe_name = "Klient"
    
    # Use orderId if provided, otherwise generate timestamp-based ID
    order_id = request.orderId if request.orderId else offer_number
    filename = f"BALIA_{safe_name}_{order_id}.pdf"
    
    return StreamingResponse(
        io.BytesIO(pdf_data),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ============================================================
# WEB ORDERS - Public API for embedded calculator
# ============================================================

from models.balia import WebOrder

@router.get("/public/prices")
async def get_public_prices():
    """Get prices for public calculator (no auth required) - with caching"""
    global _prices_cache
    
    # Check cache first
    current_time = time.time()
    if _prices_cache["data"] is not None and current_time < _prices_cache["expires"]:
        return _prices_cache["data"]
    
    prices = await db.prices.find_one({"_id": "default"})
    if not prices:
        return default_balia_prices
    
    prices.pop('_id', None)
    
    # Ensure critical fields are arrays
    if not isinstance(prices.get('models'), list):
        prices['models'] = default_balia_prices.get('models', [])
    if not isinstance(prices.get('categories'), list):
        prices['categories'] = default_balia_prices.get('categories', [])
    
    # Update cache
    _prices_cache = {"data": prices, "expires": current_time + CACHE_TTL}
    
    return prices


@router.post("/public/web-order")
async def create_web_order(order: WebOrder):
    """Create order from public website (no auth required)"""
    order_dict = order.model_dump()
    order_dict['source'] = 'website'
    await db.web_orders.insert_one(order_dict)
    
    # Send Telegram notification with PDF for web order
    try:
        from models.balia import PDFRequest
        pdf_request = PDFRequest(
            orderId=order_dict.get('id'),
            fullName=order_dict.get('fullName', ''),
            phoneNumber=order_dict.get('phoneNumber', ''),
            fullAddress=order_dict.get('fullAddress', order_dict.get('address', '')),
            orderDate=order_dict.get('orderDate', order_dict.get('createdAt', datetime.now(timezone.utc).isoformat())),
            email=order_dict.get('email'),
            notes=order_dict.get('notes'),
            modelName=order_dict.get('modelName'),
            heaterType=order_dict.get('heaterType'),
            modelSpecs=order_dict.get('modelSpecs'),
            selectedOptions=order_dict.get('selectedOptions', []),
            total=order_dict.get('total'),
            currency=order_dict.get('currency', 'PLN'),
            currencySymbol=order_dict.get('currencySymbol', 'zł')
        )
        pdf_data = await generate_pdf_bytes(pdf_request)
        await notify_new_order(order_dict, order_type='balia', is_web_order=True, pdf_data=pdf_data)
    except Exception as e:
        logger.warning(f"Failed to send Telegram notification with PDF for web order: {e}")
        # Fallback: try to send without PDF
        try:
            await notify_new_order(order_dict, order_type='balia', is_web_order=True)
        except:
            pass
    
    # Return success without sensitive data
    return {
        "success": True,
        "orderId": order.id,
        "message": "Zamówienie zostało przyjęte. Skontaktujemy się wkrótce."
    }


@router.get("/web-orders")
async def get_web_orders():
    """Get all web orders (requires auth)"""
    orders = await db.web_orders.find({}, {"_id": 0}).sort("createdAt", -1).to_list(1000)
    return orders


@router.get("/web-orders/new-count")
async def get_new_web_orders_count():
    """Get count of new (unprocessed) web orders for notifications"""
    count = await db.web_orders.count_documents({"status": "new"})
    return {"count": count}


@router.get("/web-orders/{order_id}")
async def get_web_order(order_id: str):
    """Get single web order by ID"""
    order = await db.web_orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@router.put("/web-orders/{order_id}")
async def update_web_order(order_id: str, updates: dict):
    """Update web order (status, notes, etc.)"""
    # Add timestamp for status changes
    if 'status' in updates and updates['status'] != 'new':
        updates['processedAt'] = datetime.now(timezone.utc).isoformat()
    
    result = await db.web_orders.update_one(
        {"id": order_id},
        {"$set": updates}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    updated = await db.web_orders.find_one({"id": order_id}, {"_id": 0})
    return updated


@router.delete("/web-orders/{order_id}")
async def delete_web_order(order_id: str):
    """Delete web order"""
    result = await db.web_orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"message": "Order deleted successfully"}


@router.post("/web-orders/{order_id}/transfer-to-main")
async def transfer_web_order_to_main(order_id: str, updates: dict = None):
    """Transfer web order to main orders list"""
    # Get web order
    web_order = await db.web_orders.find_one({"id": order_id}, {"_id": 0})
    if not web_order:
        raise HTTPException(status_code=404, detail="Web order not found")
    
    # Apply any updates (from editing in calculator)
    if updates:
        web_order.update(updates)
    
    # Create new order ID for main list
    new_id = f"WMB-{datetime.now(timezone.utc).strftime('%d-%m-%Y-%H%M%S')}"
    
    # Convert to main order format
    main_order = {
        "id": new_id,
        "fullName": web_order.get('customerName', ''),
        "phoneNumber": web_order.get('customerPhone', ''),
        "fullAddress": web_order.get('fullAddress', updates.get('fullAddress', '') if updates else ''),
        "orderDate": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "modelId": web_order.get('modelId'),
        "modelName": web_order.get('modelName'),
        "modelPrice": web_order.get('modelPrice', 0),
        "modelImageUrl": web_order.get('modelImageUrl'),
        "heaterType": web_order.get('heaterType') or web_order.get('heaterVariantType'),
        "heaterTypeName": web_order.get('heaterTypeName'),
        "selectedHeaterVariantId": web_order.get('selectedHeaterVariantId'),
        "selections": web_order.get('selections', {}),
        "selectedOptions": web_order.get('selectedOptions', []),
        "subtotal": web_order.get('subtotal', 0),
        "total": web_order.get('total', 0),
        "currency": web_order.get('currency', 'PLN'),
        "notes": web_order.get('customerComment', ''),
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "source": "web",  # Mark as from web
        "originalWebOrderId": order_id,  # Reference to original web order
    }
    
    # Apply additional updates if provided
    if updates:
        for key in ['fullName', 'phoneNumber', 'fullAddress', 'modelId', 'modelName', 
                    'modelPrice', 'heaterType', 'heaterTypeName', 'selections', 
                    'selectedOptions', 'subtotal', 'total', 'notes', 'discountPercent']:
            if key in updates:
                main_order[key] = updates[key]
    
    # Insert into main orders
    await db.orders.insert_one(main_order)
    
    # Update web order status to 'transferred'
    await db.web_orders.update_one(
        {"id": order_id},
        {"$set": {"status": "transferred", "transferredAt": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {
        "success": True,
        "mainOrderId": main_order_id,
        "webOrderId": order_id,
        "message": "Order transferred to main list"
    }


# ==================== TELEGRAM SETTINGS ====================

from services.telegram_service import test_telegram_connection, get_telegram_config

@router.get("/telegram/settings")
async def get_telegram_settings():
    """Get current Telegram notification settings from DB or env"""
    # First try database
    db_settings = await db.settings.find_one({"type": "telegram_settings"})
    if db_settings:
        return {
            "bot_token": db_settings.get('bot_token', '')[:20] + "..." if db_settings.get('bot_token') else "",
            "chat_id": db_settings.get('chat_id', ''),
            "enabled": db_settings.get('enabled', True),
            "bot_token_set": bool(db_settings.get('bot_token')),
            "chat_id_set": bool(db_settings.get('chat_id'))
        }
    
    # Fallback to env
    config = get_telegram_config()
    return {
        "bot_token": config['bot_token'][:20] + "..." if config['bot_token'] else "",
        "chat_id": config['chat_id'],
        "enabled": config['enabled'],
        "bot_token_set": bool(config['bot_token']),
        "chat_id_set": bool(config['chat_id'])
    }


@router.post("/telegram/test")
async def test_telegram_settings(data: dict):
    """Test Telegram connection with provided credentials"""
    bot_token = data.get('bot_token', '')
    chat_id = data.get('chat_id', '')
    
    if not bot_token or not chat_id:
        return {"success": False, "error": "Bot token and chat ID are required"}
    
    result = await test_telegram_connection(bot_token, chat_id)
    return result


@router.post("/telegram/settings")
async def update_telegram_settings(data: dict):
    """Update Telegram settings in database and environment"""
    import os
    
    # Save to database for persistence
    settings_data = {
        "type": "telegram_settings",
        "bot_token": data.get('bot_token', ''),
        "chat_id": str(data.get('chat_id', '')),
        "enabled": data.get('enabled', True)
    }
    
    await db.settings.update_one(
        {"type": "telegram_settings"},
        {"$set": settings_data},
        upsert=True
    )
    
    # Also update environment variables for current process
    if 'bot_token' in data and data['bot_token']:
        os.environ['TELEGRAM_BOT_TOKEN'] = data['bot_token']
    
    if 'chat_id' in data and data['chat_id']:
        os.environ['TELEGRAM_CHAT_ID'] = str(data['chat_id'])
    
    if 'enabled' in data:
        os.environ['TELEGRAM_NOTIFICATIONS_ENABLED'] = 'true' if data['enabled'] else 'false'
    
    return {"success": True, "message": "Settings saved to database"}
