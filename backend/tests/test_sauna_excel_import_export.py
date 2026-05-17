"""Test sauna Excel/CSV export and import (dry-run / commit) endpoints."""
import io
import os
import pytest
import requests
import openpyxl  # type: ignore

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://margin-popup-next.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"


@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{API}/auth/login", json={"username": "admin", "password": "admin123"}, timeout=30)
    if r.status_code != 200:
        pytest.skip(f"Admin login failed: {r.status_code} {r.text}")
    return r.json().get("access_token") or r.json().get("token")


@pytest.fixture(scope="module")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


@pytest.fixture(scope="module")
def dealer_id(admin_headers):
    r = requests.get(f"{API}/admin/dealers", headers=admin_headers, timeout=30)
    assert r.status_code == 200, r.text
    data = r.json()
    dealers = data.get("dealers") if isinstance(data, dict) else data
    test = next((d for d in dealers if d.get("username") == "testdealer"), None)
    if not test:
        pytest.skip("testdealer not present")
    return test["id"]


# -------- Auth --------
class TestAuth:
    def test_export_requires_auth(self):
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx", timeout=30)
        assert r.status_code in (401, 403), f"Expected 401/403, got {r.status_code}"

    def test_dry_run_requires_auth(self):
        r = requests.post(f"{API}/sauna/prices/import/dry-run",
                          files={"file": ("a.csv", b"type,id\n", "text/csv")}, timeout=30)
        assert r.status_code in (401, 403)

    def test_commit_requires_auth(self):
        r = requests.post(f"{API}/sauna/prices/import/commit",
                          files={"file": ("a.csv", b"type,id\n", "text/csv")}, timeout=30)
        assert r.status_code in (401, 403)


# -------- Export --------
class TestExport:
    def test_export_xlsx(self, admin_headers):
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx", headers=admin_headers, timeout=60)
        assert r.status_code == 200, r.text
        assert "spreadsheet" in r.headers.get("content-type", "")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "Prices" in wb.sheetnames
        ws = wb["Prices"]
        headers = [c.value for c in ws[1]]
        assert headers == ["type", "id", "parentId", "category", "name",
                           "price", "costPrice", "description", "isActive", "imageUrl"]
        assert ws.max_row > 10  # should be ~80 rows
        # Check distinct types present
        types = {ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)}
        assert "model" in types
        assert "model_variant" in types or "option" in types

    def test_export_csv(self, admin_headers):
        r = requests.get(f"{API}/sauna/prices/export?format=csv", headers=admin_headers, timeout=60)
        assert r.status_code == 200, r.text
        assert r.content[:3] == b"\xef\xbb\xbf"  # UTF-8 BOM
        text = r.content.decode("utf-8-sig")
        lines = text.splitlines()
        assert lines[0].startswith("type,id,parentId,category,name")
        assert "dealerPrice" not in lines[0]
        assert len(lines) > 10

    def test_export_xlsx_with_dealer(self, admin_headers, dealer_id):
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                         headers=admin_headers, timeout=60)
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        headers = [c.value for c in wb["Prices"][1]]
        assert "dealerPrice" in headers
        assert len(headers) == 11

    def test_export_xlsx_invalid_dealer(self, admin_headers):
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId=nonexistent_abc123",
                         headers=admin_headers, timeout=30)
        assert r.status_code == 404

    def test_export_invalid_format(self, admin_headers):
        r = requests.get(f"{API}/sauna/prices/export?format=pdf", headers=admin_headers, timeout=30)
        assert r.status_code in (400, 422)


# -------- Dry-Run --------
class TestDryRun:
    def test_dry_run_unchanged(self, admin_headers):
        # Export then immediately dry-run with same file → should be all unchanged
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx", headers=admin_headers, timeout=60)
        assert r.status_code == 200
        files = {"file": ("sauna_prices.xlsx", r.content,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        dr = requests.post(f"{API}/sauna/prices/import/dry-run",
                           headers=admin_headers, files=files, timeout=60)
        assert dr.status_code == 200, dr.text
        body = dr.json()
        s = body["summary"]
        assert s["errors"] == 0, f"unexpected errors: {body}"
        assert s["modified"] == 0, f"unexpected modifications: {s}"
        assert s["added"] == 0
        assert s["unchanged"] > 0
        assert "rows" in body and isinstance(body["rows"], list)

    def test_dry_run_modified_price(self, admin_headers):
        # Modify a model price in the file → expect 'modified' status with price diff
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx", headers=admin_headers, timeout=60)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Prices"]
        target_row = None
        original_price = None
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=1).value == "model":
                target_row = i
                original_price = int(ws.cell(row=i, column=6).value or 0)
                ws.cell(row=i, column=6, value=original_price + 777)
                break
        assert target_row is not None
        buf = io.BytesIO()
        wb.save(buf)
        files = {"file": ("modified.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        dr = requests.post(f"{API}/sauna/prices/import/dry-run",
                           headers=admin_headers, files=files, timeout=60)
        assert dr.status_code == 200, dr.text
        body = dr.json()
        assert body["summary"]["modified"] >= 1
        modified_rows = [r for r in body["rows"] if r["status"] == "modified"]
        assert any("price" in (r.get("diff") or {}) for r in modified_rows)
        sample = next(r for r in modified_rows if "price" in (r.get("diff") or {}))
        assert sample["diff"]["price"]["old"] == original_price
        assert sample["diff"]["price"]["new"] == original_price + 777

    def test_dry_run_empty_file(self, admin_headers):
        files = {"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{API}/sauna/prices/import/dry-run",
                          headers=admin_headers, files=files, timeout=30)
        assert r.status_code == 400

    def test_dry_run_invalid_file(self, admin_headers):
        files = {"file": ("readme.txt", b"hello world", "text/plain")}
        r = requests.post(f"{API}/sauna/prices/import/dry-run",
                          headers=admin_headers, files=files, timeout=30)
        assert r.status_code == 400


# -------- Commit (base + dealer overrides) --------
class TestCommit:
    def test_commit_no_change_then_modify_then_revert(self, admin_headers):
        # 1) baseline GET
        r0 = requests.get(f"{API}/sauna/prices", timeout=30)
        assert r0.status_code == 200
        models = r0.json().get("models", [])
        assert models, "no models in DB"
        target_model_id = models[0]["id"]
        original_price = int(models[0].get("basePrice") or 0)

        # 2) Export and bump price by +500
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx", headers=admin_headers, timeout=60)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Prices"]
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=1).value == "model" and ws.cell(row=i, column=2).value == target_model_id:
                ws.cell(row=i, column=6, value=original_price + 500)
                break
        buf = io.BytesIO(); wb.save(buf)

        # 3) Commit
        files = {"file": ("upd.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        cr = requests.post(f"{API}/sauna/prices/import/commit",
                           headers=admin_headers, files=files, timeout=60)
        assert cr.status_code == 200, cr.text
        body = cr.json()
        assert body["ok"] is True
        assert body["summary"]["modified"] >= 1

        # 4) Verify GET reflects new price
        r2 = requests.get(f"{API}/sauna/prices", timeout=30)
        models2 = r2.json().get("models", [])
        new_price = next((m["basePrice"] for m in models2 if m["id"] == target_model_id), None)
        assert new_price == original_price + 500

        # 5) Revert: re-export and commit back
        rr = requests.get(f"{API}/sauna/prices/export?format=xlsx", headers=admin_headers, timeout=60)
        wb2 = openpyxl.load_workbook(io.BytesIO(rr.content))
        ws2 = wb2["Prices"]
        for i in range(2, ws2.max_row + 1):
            if ws2.cell(row=i, column=1).value == "model" and ws2.cell(row=i, column=2).value == target_model_id:
                ws2.cell(row=i, column=6, value=original_price)
                break
        buf2 = io.BytesIO(); wb2.save(buf2)
        cr2 = requests.post(f"{API}/sauna/prices/import/commit",
                            headers=admin_headers,
                            files={"file": ("revert.xlsx", buf2.getvalue(),
                                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                            timeout=60)
        assert cr2.status_code == 200

    def test_commit_dealer_override(self, admin_headers, dealer_id):
        # Export for dealer; pick a model row; set dealerPrice
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                         headers=admin_headers, timeout=60)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Prices"]
        headers = [c.value for c in ws[1]]
        dp_col = headers.index("dealerPrice") + 1
        target_row = None; target_id = None
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=1).value == "model":
                target_row = i
                target_id = ws.cell(row=i, column=2).value
                ws.cell(row=i, column=dp_col, value=99999)
                break
        # Leave other rows' dealerPrice as-is (empty / current)
        buf = io.BytesIO(); wb.save(buf)
        files = {"file": ("dp.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        cr = requests.post(f"{API}/sauna/prices/import/commit",
                           headers=admin_headers,
                           data={"dealerId": dealer_id},
                           files=files, timeout=60)
        assert cr.status_code == 200, cr.text
        body = cr.json()
        assert body["overridesUpserted"] >= 1

        # Re-export and verify dealerPrice persisted
        r2 = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                          headers=admin_headers, timeout=60)
        wb2 = openpyxl.load_workbook(io.BytesIO(r2.content))
        ws2 = wb2["Prices"]
        for i in range(2, ws2.max_row + 1):
            if ws2.cell(row=i, column=1).value == "model" and ws2.cell(row=i, column=2).value == target_id:
                assert int(ws2.cell(row=i, column=dp_col).value or 0) == 99999
                break

    def test_commit_empty_dealerprice_is_noop(self, admin_headers, dealer_id):
        # Export for dealer, blank out all dealerPrice cells, commit → overridesUpserted == 0
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                         headers=admin_headers, timeout=60)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Prices"]
        headers = [c.value for c in ws[1]]
        dp_col = headers.index("dealerPrice") + 1
        for i in range(2, ws.max_row + 1):
            ws.cell(row=i, column=dp_col, value="")
        buf = io.BytesIO(); wb.save(buf)
        cr = requests.post(f"{API}/sauna/prices/import/commit",
                           headers=admin_headers,
                           data={"dealerId": dealer_id},
                           files={"file": ("blank.xlsx", buf.getvalue(),
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                           timeout=60)
        assert cr.status_code == 200, cr.text
        assert cr.json()["overridesUpserted"] == 0
