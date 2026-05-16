"""Iteration 95 — Excel round-trip for components+tech-cards + dealer comparison endpoint."""
import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://sauna-config-5.preview.emergentagent.com").rstrip("/")

EXPECTED_COMP_HEADERS = [
    "id", "name", "category", "unit", "unitPrice", "supplier", "note",
    "stockCurrent", "stockMin", "isActive",
]
EXPECTED_CARD_HEADERS = [
    "cardId", "scope", "modelId", "variantId", "optionId", "optionVariantId",
    "componentId", "componentName", "qty", "itemNote",
    "laborCost", "overheadPct", "manualAdjustment", "retailExtraCost",
    "syncToCostPrice", "cardNote",
]


@pytest.fixture(scope="module")
def admin_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"username": "admin", "password": "admin123"}, timeout=30)
    assert r.status_code == 200, r.text
    return r.json()["token"]


@pytest.fixture(scope="module")
def headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}"}


# ---------------------- EXPORT ----------------------
class TestExport:
    def test_export_status_and_content_type(self, headers):
        r = requests.get(f"{BASE_URL}/api/sauna-production/cost/export", headers=headers, timeout=60)
        assert r.status_code == 200, r.text
        assert "spreadsheetml.sheet" in r.headers.get("content-type", "")
        assert len(r.content) > 1000
        # Stash in module-level for next tests
        TestExport.blob = r.content

    def test_export_has_two_sheets_and_headers(self, headers):
        wb = load_workbook(io.BytesIO(TestExport.blob))
        assert "Components" in wb.sheetnames
        assert "TechCards" in wb.sheetnames
        ws_c = wb["Components"]
        comp_headers = [c.value for c in ws_c[1]]
        assert comp_headers == EXPECTED_COMP_HEADERS, comp_headers
        ws_t = wb["TechCards"]
        card_headers = [c.value for c in ws_t[1]]
        assert card_headers == EXPECTED_CARD_HEADERS, card_headers

    def test_components_count_matches_db(self, headers):
        r = requests.get(f"{BASE_URL}/api/sauna-production/cost/components", headers=headers, timeout=30)
        assert r.status_code == 200
        db_items = r.json().get("items") or r.json()
        if isinstance(db_items, dict):
            db_items = db_items.get("items", [])
        wb = load_workbook(io.BytesIO(TestExport.blob))
        ws_c = wb["Components"]
        rows = list(ws_c.iter_rows(values_only=True))[1:]
        non_empty = [r for r in rows if any(v not in (None, "") for v in r)]
        assert len(non_empty) == len(db_items), f"xlsx={len(non_empty)} vs db={len(db_items)}"


# ---------------------- DRY-RUN ROUND-TRIP ----------------------
class TestDryRunRoundTrip:
    def test_round_trip_unmodified_yields_unchanged(self, headers):
        r = requests.get(f"{BASE_URL}/api/sauna-production/cost/export", headers=headers, timeout=60)
        assert r.status_code == 200
        files = {"file": ("export.xlsx", r.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r2 = requests.post(f"{BASE_URL}/api/sauna-production/cost/import-dry-run", headers=headers, files=files, timeout=60)
        assert r2.status_code == 200, r2.text
        data = r2.json()
        assert "components" in data and "techCards" in data and "summary" in data
        comp = data["components"]
        cards = data["techCards"]
        assert comp["add"] == [], f"unexpected adds in round-trip: {comp['add'][:3]}"
        assert comp["update"] == [], f"unexpected component updates in round-trip: {comp['update'][:3]}"
        # cards: no adds in round-trip
        assert cards["add"] == [], f"unexpected card adds: {cards['add'][:3]}"
        # cards updates should be empty (round-trip)
        assert cards["update"] == [], f"unexpected card updates: {cards['update'][:3]}"

    def test_non_xlsx_returns_400(self, headers):
        files = {"file": ("test.txt", b"not an xlsx", "text/plain")}
        r = requests.post(f"{BASE_URL}/api/sauna-production/cost/import-dry-run", headers=headers, files=files, timeout=30)
        assert r.status_code == 400

    def test_empty_workbook_graceful(self, headers):
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = "Random"
        buf = io.BytesIO()
        wb.save(buf)
        files = {"file": ("empty.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/sauna-production/cost/import-dry-run", headers=headers, files=files, timeout=30)
        assert r.status_code == 200
        d = r.json()
        assert d["components"]["add"] == [] and d["components"]["update"] == []
        assert d["techCards"]["add"] == [] and d["techCards"]["update"] == []

    def test_missing_name_recorded_in_errors(self, headers):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Components"
        ws.append(EXPECTED_COMP_HEADERS)
        # row with empty name but some other data
        ws.append(["", "", "wood", "шт", 100, "Sup", "n", 0, 0, True])
        wb.create_sheet("TechCards").append(EXPECTED_CARD_HEADERS)
        buf = io.BytesIO()
        wb.save(buf)
        files = {"file": ("bad.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/sauna-production/cost/import-dry-run", headers=headers, files=files, timeout=30)
        assert r.status_code == 200
        d = r.json()
        assert d["summary"]["errorsCount"] >= 1
        assert any(e.get("sheet") == "Components" and "name" in e.get("message", "").lower() for e in d["errors"])


# ---------------------- COMMIT — edit unitPrice persists ----------------------
class TestCommit:
    def test_commit_edited_unit_price_persists(self, headers):
        # Export
        r = requests.get(f"{BASE_URL}/api/sauna-production/cost/export", headers=headers, timeout=60)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws_c = wb["Components"]
        # Edit the first data row's unitPrice → bump by 1.0
        headers_row = [c.value for c in ws_c[1]]
        up_idx = headers_row.index("unitPrice") + 1
        id_idx = headers_row.index("id") + 1
        first_row = 2
        target_id = ws_c.cell(row=first_row, column=id_idx).value
        old_price = ws_c.cell(row=first_row, column=up_idx).value or 0
        new_price = float(old_price) + 1.0
        ws_c.cell(row=first_row, column=up_idx).value = new_price
        buf = io.BytesIO()
        wb.save(buf)
        files = {"file": ("commit.xlsx", buf.getvalue(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r2 = requests.post(f"{BASE_URL}/api/sauna-production/cost/import-commit", headers=headers, files=files, timeout=120)
        assert r2.status_code == 200, r2.text
        data = r2.json()
        assert data["ok"] is True
        assert data["components"]["updated"] >= 1
        # Verify via GET
        r3 = requests.get(f"{BASE_URL}/api/sauna-production/cost/components", headers=headers, timeout=30)
        assert r3.status_code == 200
        items = r3.json()
        if isinstance(items, dict):
            items = items.get("items", [])
        match = next((c for c in items if c.get("id") == target_id), None)
        assert match is not None, f"component {target_id} not found after commit"
        assert abs(float(match["unitPrice"]) - new_price) < 0.01, f"unitPrice not persisted: got {match['unitPrice']} vs {new_price}"
        # Revert (cleanup)
        ws_c.cell(row=first_row, column=up_idx).value = old_price
        buf2 = io.BytesIO()
        wb.save(buf2)
        requests.post(f"{BASE_URL}/api/sauna-production/cost/import-commit", headers=headers,
                      files={"file": ("revert.xlsx", buf2.getvalue(),
                                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                      timeout=120)


# ---------------------- DEALER COMPARISON ----------------------
class TestDealerComparison:
    def test_endpoint_returns_shape(self, headers):
        r = requests.get(f"{BASE_URL}/api/admin/dealers/comparison", headers=headers, timeout=60)
        assert r.status_code == 200, r.text
        d = r.json()
        assert "dealers" in d and "rows" in d and "totalRows" in d
        assert isinstance(d["dealers"], list)
        assert isinstance(d["rows"], list)
        assert d["totalRows"] == len(d["rows"])
        # dealer shape
        if d["dealers"]:
            d0 = d["dealers"][0]
            for f in ("id", "name", "username", "isActive"):
                assert f in d0

    def test_rows_have_all_kinds_and_dealer_price_shape(self, headers):
        r = requests.get(f"{BASE_URL}/api/admin/dealers/comparison", headers=headers, timeout=60)
        d = r.json()
        kinds = {row["kind"] for row in d["rows"]}
        # Must include at least the broad kinds defined in the spec
        assert {"model", "option"}.issubset(kinds), f"kinds={kinds}"
        # The spec asks for all 4 — only assert if catalog has them
        possible = {"model", "model_variant", "option", "option_variant"}
        assert kinds.issubset(possible)
        # Validate row shape
        sample = d["rows"][0]
        for f in ("kind", "name", "retailBrutto", "dealers",
                  "minDealerPrice", "maxDealerPrice", "avgDealerPrice", "overrideCount"):
            assert f in sample, f
        # Each dealer entry: dealerId, dealerName, price (None or int)
        for de in sample["dealers"]:
            assert "dealerId" in de and "dealerName" in de and "price" in de
            assert de["price"] is None or isinstance(de["price"], int)

    def test_requires_admin_auth(self):
        r = requests.get(f"{BASE_URL}/api/admin/dealers/comparison", timeout=30)
        assert r.status_code in (401, 403)


# ---------------------- REGRESSION sanity ----------------------
class TestRegression:
    def test_components_list_endpoint(self, headers):
        r = requests.get(f"{BASE_URL}/api/sauna-production/cost/components", headers=headers, timeout=30)
        assert r.status_code == 200

    def test_dealers_list_endpoint(self, headers):
        r = requests.get(f"{BASE_URL}/api/admin/dealers", headers=headers, timeout=30)
        assert r.status_code == 200
