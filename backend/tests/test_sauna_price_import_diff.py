"""Tests for the new per-entry history diff endpoint:
   GET /api/sauna/prices/import/history/{id}/diff
"""
import io
import os
import pytest
import requests
import openpyxl  # type: ignore

from motor.motor_asyncio import AsyncIOMotorClient
import asyncio

def _load_env(path):
    try:
        with open(path) as fh:
            for ln in fh:
                ln = ln.strip()
                if not ln or ln.startswith("#") or "=" not in ln:
                    continue
                k, v = ln.split("=", 1)
                v = v.strip().strip('"').strip("'")
                os.environ.setdefault(k.strip(), v)
    except FileNotFoundError:
        pass

_load_env("/app/frontend/.env")
_load_env("/app/backend/.env")

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
API = f"{BASE_URL}/api"
MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME")


# ----------------------------- fixtures -----------------------------
@pytest.fixture(scope="module")
def admin_headers():
    r = requests.post(f"{API}/auth/login",
                      json={"username": "admin", "password": "admin123"},
                      timeout=30)
    if r.status_code != 200:
        pytest.skip(f"Admin login failed: {r.status_code} {r.text}")
    tok = r.json().get("access_token") or r.json().get("token")
    return {"Authorization": f"Bearer {tok}"}


def _modify_first_model_price_and_cost(content, price_bump=355, cost_bump=999):
    """Bump price (col 6) and bump costPrice (col 7) of the FIRST `model` row."""
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Prices"]
    target_id = None
    old_price = None
    old_cost = None
    new_cost = None
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == "model":
            target_id = ws.cell(row=i, column=2).value
            old_price = int(ws.cell(row=i, column=6).value or 0)
            old_cost = int(ws.cell(row=i, column=7).value or 0)
            new_cost = old_cost + cost_bump
            ws.cell(row=i, column=6, value=old_price + price_bump)
            ws.cell(row=i, column=7, value=new_cost)
            break
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), target_id, old_price, old_cost, new_cost


def _commit_modify_first_model(admin_headers):
    r = requests.get(f"{API}/sauna/prices/export?format=xlsx", headers=admin_headers, timeout=60)
    assert r.status_code == 200
    payload, target_id, old_price, old_cost, new_cost = _modify_first_model_price_and_cost(r.content)
    files = {"file": ("diff_test.xlsx", payload,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    cr = requests.post(f"{API}/sauna/prices/import/commit",
                       headers=admin_headers, files=files, timeout=60)
    assert cr.status_code == 200, cr.text
    return cr.json(), target_id, old_price, old_cost, new_cost


# ---------------------------- AUTH -----------------------------------
class TestDiffAuth:
    def test_diff_requires_auth(self):
        r = requests.get(f"{API}/sauna/prices/import/history/anything/diff", timeout=30)
        assert r.status_code in (401, 403)


# ------------------------ Fresh commit diff --------------------------
class TestFreshDiff:
    def test_fresh_commit_diff_modified_one_row(self, admin_headers):
        commit, target_id, old_price, old_cost, new_cost = _commit_modify_first_model(admin_headers)
        hid = commit["historyId"]
        new_price = old_price + 355

        d = requests.get(f"{API}/sauna/prices/import/history/{hid}/diff",
                         headers=admin_headers, timeout=30)
        assert d.status_code == 200, d.text
        body = d.json()

        # top-level metadata
        assert body["historyId"] == hid
        assert body["isFallback"] is False
        assert body["filename"] == "diff_test.xlsx"
        assert body["adminUsername"] in ("admin",)
        assert body["timestamp"]
        assert body["dealerId"] is None
        assert body["rolledBack"] is False

        # summary
        s = body["summary"]
        for k in ("added", "modified", "removed", "unchanged", "marginAlerts"):
            assert k in s, f"missing {k} in summary: {s}"
        assert s["modified"] == 1
        assert s["added"] == 0
        assert s["removed"] == 0
        assert s["unchanged"] > 0

        # locate the modified row
        rows = body["rows"]
        mod = [r for r in rows if r["status"] == "modified"]
        assert len(mod) == 1, f"expected exactly 1 modified, got {len(mod)}: {[r.get('id') for r in mod]}"
        row = mod[0]
        assert row["type"] == "model"
        assert row["id"] == target_id

        # diff shape: price + costPrice changed
        d_block = row["diff"]
        assert "price" in d_block
        assert d_block["price"]["old"] == old_price
        assert d_block["price"]["new"] == new_price
        assert "costPrice" in d_block
        assert d_block["costPrice"]["old"] == old_cost
        assert d_block["costPrice"]["new"] == new_cost

        # margin block populated
        m = row["margin"]
        assert m["newAmount"] == new_price - new_cost
        # newPct should be a number
        assert isinstance(m["newPct"], (int, float))

    def test_unchanged_rows_present_for_other_entities(self, admin_headers):
        commit, _, _, _, _ = _commit_modify_first_model(admin_headers)
        hid = commit["historyId"]
        body = requests.get(f"{API}/sauna/prices/import/history/{hid}/diff",
                            headers=admin_headers, timeout=30).json()
        unchanged = [r for r in body["rows"] if r["status"] == "unchanged"]
        assert len(unchanged) >= 1
        # status counts agree with summary
        assert body["summary"]["unchanged"] == len(unchanged)
        assert body["summary"]["modified"] == len([r for r in body["rows"] if r["status"] == "modified"])


# ---------------------- List endpoint excludes snapshots ----------------
class TestHistoryListNoSnapshots:
    def test_list_payload_excludes_snapshot_blobs(self, admin_headers):
        # create at least one entry to ensure the list is non-empty
        _commit_modify_first_model(admin_headers)
        r = requests.get(f"{API}/sauna/prices/import/history?limit=5",
                         headers=admin_headers, timeout=30)
        assert r.status_code == 200, r.text
        items = r.json().get("items") or []
        assert len(items) >= 1
        for it in items:
            for key in ("snapshotPrices", "snapshotOverrides",
                        "snapshotAfterPrices", "snapshotAfterOverrides"):
                assert key not in it, f"{key} should NOT be present in list payload: {it.keys()}"


# ---------------------- 404 for unknown id --------------------------------
class TestDiff404:
    def test_diff_unknown_id_returns_404(self, admin_headers):
        r = requests.get(f"{API}/sauna/prices/import/history/does-not-exist-uuid/diff",
                         headers=admin_headers, timeout=30)
        assert r.status_code == 404


# ---------------------- Legacy fallback path ------------------------------
class TestFallback:
    def test_fallback_when_snapshot_after_missing(self, admin_headers):
        """Simulate a legacy entry by clearing snapshotAfterPrices in DB, then
        verify the diff endpoint falls back to live state and sets isFallback=true."""
        if not MONGO_URL or not DB_NAME:
            pytest.skip("MONGO_URL/DB_NAME not set in env")

        commit, target_id, old_price, old_cost, new_cost = _commit_modify_first_model(admin_headers)
        hid = commit["historyId"]

        async def _strip_snapshot_after():
            client = AsyncIOMotorClient(MONGO_URL)
            try:
                db = client[DB_NAME]
                res = await db.sauna_price_import_history.update_one(
                    {"id": hid},
                    {"$unset": {"snapshotAfterPrices": "",
                                "snapshotAfterOverrides": ""}},
                )
                return res.modified_count
            finally:
                client.close()

        modified = asyncio.get_event_loop().run_until_complete(_strip_snapshot_after()) \
            if not asyncio.get_event_loop().is_running() else asyncio.new_event_loop().run_until_complete(_strip_snapshot_after())
        assert modified == 1

        d = requests.get(f"{API}/sauna/prices/import/history/{hid}/diff",
                         headers=admin_headers, timeout=30)
        assert d.status_code == 200, d.text
        body = d.json()
        assert body["isFallback"] is True
        # Should still produce a summary + rows even in fallback mode
        assert "summary" in body
        assert "rows" in body
        # Since we just committed and live state == snapshotAfter, the diff
        # before↔live should still show the same `modified` row.
        mod = [r for r in body["rows"] if r["status"] == "modified"]
        ids = [r["id"] for r in mod]
        assert target_id in ids, f"expected {target_id} in modified rows, got {ids}"


# ---------------------- 'removed' status ----------------------------------
class TestRemovedStatus:
    def test_removed_status_when_entity_disappears_between_snapshots(self, admin_headers):
        """Inject a fake history entry whose before-snapshot contains an extra
        model that is NOT in the after-snapshot — diff must report status='removed'."""
        if not MONGO_URL or not DB_NAME:
            pytest.skip("MONGO_URL/DB_NAME not set in env")

        import uuid
        fake_id = f"TEST_REMOVED_{uuid.uuid4().hex[:6]}"
        hid = f"TEST_HIST_{uuid.uuid4().hex[:8]}"
        before = {
            "models": [
                {"id": fake_id, "name": "Sauna to remove",
                 "basePrice": 10000, "costPrice": 4000,
                 "active": True, "variants": []}
            ],
            "categories": [], "options": [],
        }
        after = {"models": [], "categories": [], "options": []}

        async def _insert_and_cleanup():
            client = AsyncIOMotorClient(MONGO_URL)
            try:
                db = client[DB_NAME]
                from datetime import datetime, timezone
                await db.sauna_price_import_history.insert_one({
                    "id": hid,
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "adminUsername": "admin",
                    "filename": "test_removed.xlsx",
                    "dealerId": None,
                    "dealerName": "",
                    "summary": {"added": 0, "modified": 0, "unchanged": 0,
                                "errors": 0, "overrides_changed": 0},
                    "overridesUpserted": 0,
                    "totalRows": 1,
                    "snapshotPrices": before,
                    "snapshotOverrides": None,
                    "snapshotAfterPrices": after,
                    "snapshotAfterOverrides": None,
                    "rolledBack": False,
                })
            finally:
                client.close()

        async def _delete():
            client = AsyncIOMotorClient(MONGO_URL)
            try:
                db = client[DB_NAME]
                await db.sauna_price_import_history.delete_one({"id": hid})
            finally:
                client.close()

        loop = asyncio.new_event_loop()
        try:
            loop.run_until_complete(_insert_and_cleanup())

            d = requests.get(f"{API}/sauna/prices/import/history/{hid}/diff",
                             headers=admin_headers, timeout=30)
            assert d.status_code == 200, d.text
            body = d.json()
            assert body["summary"]["removed"] == 1
            assert body["summary"]["added"] == 0
            assert body["summary"]["modified"] == 0
            rem = [r for r in body["rows"] if r["status"] == "removed"]
            assert len(rem) == 1
            assert rem[0]["id"] == fake_id
            assert rem[0]["type"] == "model"
            # Removed-row diff must show price old=10000, new=None
            assert rem[0]["diff"]["price"]["old"] == 10000
            assert rem[0]["diff"]["price"]["new"] is None
        finally:
            loop.run_until_complete(_delete())
            loop.close()
