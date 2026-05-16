"""Test sauna price import HISTORY + ROLLBACK endpoints."""
import io
import os
import pytest
import requests
import openpyxl  # type: ignore

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
API = f"{BASE_URL}/api"


# ---- fixtures ----
@pytest.fixture(scope="module")
def admin_headers():
    r = requests.post(f"{API}/auth/login", json={"username": "admin", "password": "admin123"}, timeout=30)
    if r.status_code != 200:
        pytest.skip(f"Admin login failed: {r.status_code} {r.text}")
    tok = r.json().get("access_token") or r.json().get("token")
    return {"Authorization": f"Bearer {tok}"}


@pytest.fixture(scope="module")
def dealer_id(admin_headers):
    r = requests.get(f"{API}/admin/dealers", headers=admin_headers, timeout=30)
    assert r.status_code == 200, r.text
    data = r.json()
    dealers = data.get("dealers") if isinstance(data, dict) else data
    test = next((d for d in dealers if d.get("username") == "testdealer"), None)
    if not test:
        pytest.skip("testdealer not present")
    return test["id"]


def _bump_first_model(content, bump=333):
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Prices"]
    target_id = None
    original = None
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == "model":
            target_id = ws.cell(row=i, column=2).value
            original = int(ws.cell(row=i, column=6).value or 0)
            ws.cell(row=i, column=6, value=original + bump)
            break
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), target_id, original


def _commit_global_bump(admin_headers, bump=333):
    r = requests.get(f"{API}/sauna/prices/export?format=xlsx", headers=admin_headers, timeout=60)
    assert r.status_code == 200
    payload, target_id, original = _bump_first_model(r.content, bump=bump)
    files = {"file": ("h.xlsx", payload,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    cr = requests.post(f"{API}/sauna/prices/import/commit",
                       headers=admin_headers, files=files, timeout=60)
    assert cr.status_code == 200, cr.text
    return cr.json(), target_id, original


# ============================================================
# Auth checks – every history endpoint must require admin
# ============================================================
class TestHistoryAuth:
    def test_list_requires_auth(self):
        r = requests.get(f"{API}/sauna/prices/import/history", timeout=30)
        assert r.status_code in (401, 403)

    def test_rollback_requires_auth(self):
        r = requests.post(f"{API}/sauna/prices/import/history/some-id/rollback", timeout=30)
        assert r.status_code in (401, 403)

    def test_delete_requires_auth(self):
        r = requests.delete(f"{API}/sauna/prices/import/history/some-id", timeout=30)
        assert r.status_code in (401, 403)


# ============================================================
# Commit returns historyId + list endpoint
# ============================================================
class TestHistoryList:
    def test_commit_returns_history_id(self, admin_headers):
        body, target_id, original = _commit_global_bump(admin_headers, bump=111)
        assert "historyId" in body and isinstance(body["historyId"], str) and len(body["historyId"]) > 8
        # revert immediately to keep DB sane (use rollback)
        rb = requests.post(
            f"{API}/sauna/prices/import/history/{body['historyId']}/rollback",
            headers=admin_headers, timeout=30,
        )
        assert rb.status_code == 200, rb.text

    def test_list_sorted_desc_no_snapshot_blobs(self, admin_headers):
        # Make 2 commits then list
        b1, _, _ = _commit_global_bump(admin_headers, bump=11)
        b2, _, _ = _commit_global_bump(admin_headers, bump=22)
        r = requests.get(f"{API}/sauna/prices/import/history?limit=20",
                         headers=admin_headers, timeout=30)
        assert r.status_code == 200, r.text
        data = r.json()
        assert "items" in data and isinstance(data["items"], list)
        items = data["items"]
        assert len(items) >= 2
        # Snapshot blobs must be excluded from list view
        for it in items:
            assert "snapshotPrices" not in it
            assert "snapshotOverrides" not in it
            assert "summary" in it
            assert "filename" in it
            assert "adminUsername" in it
        # Newest first (b2 should appear before b1)
        ids = [it["id"] for it in items]
        assert ids.index(b2["historyId"]) < ids.index(b1["historyId"])
        # Cleanup: rollback both
        for hid in (b2["historyId"], b1["historyId"]):
            requests.post(f"{API}/sauna/prices/import/history/{hid}/rollback",
                          headers=admin_headers, timeout=30)

    def test_list_dealer_scoped(self, admin_headers, dealer_id):
        # Commit a dealer-scoped change
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                         headers=admin_headers, timeout=60)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Prices"]
        headers = [c.value for c in ws[1]]
        dp_col = headers.index("dealerPrice") + 1
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=1).value == "model":
                ws.cell(row=i, column=dp_col, value=88888)
                break
        buf = io.BytesIO(); wb.save(buf)
        cr = requests.post(f"{API}/sauna/prices/import/commit",
                           headers=admin_headers,
                           data={"dealerId": dealer_id},
                           files={"file": ("d.xlsx", buf.getvalue(),
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                           timeout=60)
        assert cr.status_code == 200, cr.text
        hid = cr.json()["historyId"]

        # Dealer-scoped list contains the entry
        rl = requests.get(f"{API}/sauna/prices/import/history?dealerId={dealer_id}",
                          headers=admin_headers, timeout=30)
        assert rl.status_code == 200
        items = rl.json()["items"]
        assert any(it["id"] == hid for it in items)
        # Each item is for this dealer only
        for it in items:
            assert it.get("dealerId") == dealer_id

        # Global list (dealerId omitted) does NOT include dealer entry
        rg = requests.get(f"{API}/sauna/prices/import/history",
                          headers=admin_headers, timeout=30)
        assert rg.status_code == 200
        global_ids = [it["id"] for it in rg.json()["items"]]
        assert hid not in global_ids

        # Cleanup
        requests.post(f"{API}/sauna/prices/import/history/{hid}/rollback",
                      headers=admin_headers, timeout=30)


# ============================================================
# Rollback flow
# ============================================================
class TestRollback:
    def test_rollback_restores_prices(self, admin_headers):
        # Baseline price
        r0 = requests.get(f"{API}/sauna/prices", timeout=30)
        assert r0.status_code == 200
        models0 = r0.json()["models"]
        # Commit a bump
        body, target_id, original = _commit_global_bump(admin_headers, bump=777)
        # Verify changed
        r1 = requests.get(f"{API}/sauna/prices", timeout=30)
        new_price = next(m["basePrice"] for m in r1.json()["models"] if m["id"] == target_id)
        assert new_price == original + 777
        # Rollback
        rb = requests.post(f"{API}/sauna/prices/import/history/{body['historyId']}/rollback",
                           headers=admin_headers, timeout=30)
        assert rb.status_code == 200, rb.text
        assert rb.json()["ok"] is True
        # Verify reverted
        r2 = requests.get(f"{API}/sauna/prices", timeout=30)
        reverted = next(m["basePrice"] for m in r2.json()["models"] if m["id"] == target_id)
        assert reverted == original

    def test_rollback_twice_returns_400(self, admin_headers):
        body, _, _ = _commit_global_bump(admin_headers, bump=55)
        hid = body["historyId"]
        r1 = requests.post(f"{API}/sauna/prices/import/history/{hid}/rollback",
                           headers=admin_headers, timeout=30)
        assert r1.status_code == 200
        r2 = requests.post(f"{API}/sauna/prices/import/history/{hid}/rollback",
                           headers=admin_headers, timeout=30)
        assert r2.status_code == 400
        body2 = r2.json()
        msg = body2.get("detail") or body2.get("message") or ""
        assert "rolled back" in str(msg).lower()

    def test_rollback_unknown_id_404(self, admin_headers):
        r = requests.post(f"{API}/sauna/prices/import/history/nonexistent-id-12345/rollback",
                          headers=admin_headers, timeout=30)
        assert r.status_code == 404

    def test_rollback_dealer_scoped_restores_overrides(self, admin_headers, dealer_id):
        # Snapshot current overrides for dealer
        # Find some model id we can use
        rprices = requests.get(f"{API}/sauna/prices", timeout=30)
        target_id = rprices.json()["models"][0]["id"]

        # Set an initial override 11111
        r = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                         headers=admin_headers, timeout=60)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Prices"]
        headers = [c.value for c in ws[1]]
        dp_col = headers.index("dealerPrice") + 1
        for i in range(2, ws.max_row + 1):
            if ws.cell(row=i, column=1).value == "model" and ws.cell(row=i, column=2).value == target_id:
                ws.cell(row=i, column=dp_col, value=11111)
                break
        buf = io.BytesIO(); wb.save(buf)
        cr0 = requests.post(f"{API}/sauna/prices/import/commit",
                            headers=admin_headers,
                            data={"dealerId": dealer_id},
                            files={"file": ("init.xlsx", buf.getvalue(),
                                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                            timeout=60)
        assert cr0.status_code == 200, cr0.text
        # Verify initial override 11111 present
        rexp = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                            headers=admin_headers, timeout=60)
        wb2 = openpyxl.load_workbook(io.BytesIO(rexp.content))
        ws2 = wb2["Prices"]
        for i in range(2, ws2.max_row + 1):
            if ws2.cell(row=i, column=1).value == "model" and ws2.cell(row=i, column=2).value == target_id:
                assert int(ws2.cell(row=i, column=dp_col).value or 0) == 11111
                break

        # Now do a 2nd commit changing override to 22222 (this commit's snapshot has 11111)
        for i in range(2, ws2.max_row + 1):
            if ws2.cell(row=i, column=1).value == "model" and ws2.cell(row=i, column=2).value == target_id:
                ws2.cell(row=i, column=dp_col, value=22222)
                break
        buf2 = io.BytesIO(); wb2.save(buf2)
        cr = requests.post(f"{API}/sauna/prices/import/commit",
                           headers=admin_headers,
                           data={"dealerId": dealer_id},
                           files={"file": ("upd.xlsx", buf2.getvalue(),
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                           timeout=60)
        assert cr.status_code == 200, cr.text
        hid = cr.json()["historyId"]

        # Verify override = 22222 in fresh export
        rexp3 = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                             headers=admin_headers, timeout=60)
        wb3 = openpyxl.load_workbook(io.BytesIO(rexp3.content))
        ws3 = wb3["Prices"]
        for i in range(2, ws3.max_row + 1):
            if ws3.cell(row=i, column=1).value == "model" and ws3.cell(row=i, column=2).value == target_id:
                assert int(ws3.cell(row=i, column=dp_col).value or 0) == 22222
                break

        # Rollback the 2nd commit → override should revert to 11111
        rb = requests.post(f"{API}/sauna/prices/import/history/{hid}/rollback",
                           headers=admin_headers, timeout=30)
        assert rb.status_code == 200, rb.text
        body = rb.json()
        assert body["dealerId"] == dealer_id
        assert body["ok"] is True
        # restoredOverrides reports count present in snapshot
        assert isinstance(body.get("restoredOverrides"), int)

        rexp4 = requests.get(f"{API}/sauna/prices/export?format=xlsx&dealerId={dealer_id}",
                             headers=admin_headers, timeout=60)
        wb4 = openpyxl.load_workbook(io.BytesIO(rexp4.content))
        ws4 = wb4["Prices"]
        for i in range(2, ws4.max_row + 1):
            if ws4.cell(row=i, column=1).value == "model" and ws4.cell(row=i, column=2).value == target_id:
                assert int(ws4.cell(row=i, column=dp_col).value or 0) == 11111, "override was not reverted to snapshot value"
                break

        # Cleanup: rollback the first commit too
        # Find the first history entry id
        rl = requests.get(f"{API}/sauna/prices/import/history?dealerId={dealer_id}",
                          headers=admin_headers, timeout=30)
        first_entry = next(it for it in rl.json()["items"] if it["id"] == cr0.json()["historyId"])
        if not first_entry.get("rolledBack"):
            requests.post(
                f"{API}/sauna/prices/import/history/{first_entry['id']}/rollback",
                headers=admin_headers, timeout=30,
            )


# ============================================================
# DELETE endpoint
# ============================================================
class TestHistoryDelete:
    def test_delete_history_entry(self, admin_headers):
        body, _, _ = _commit_global_bump(admin_headers, bump=44)
        hid = body["historyId"]
        # Rollback first to clean up the price change side effect
        requests.post(f"{API}/sauna/prices/import/history/{hid}/rollback",
                      headers=admin_headers, timeout=30)
        # Delete
        d = requests.delete(f"{API}/sauna/prices/import/history/{hid}",
                            headers=admin_headers, timeout=30)
        assert d.status_code == 200, d.text
        assert d.json().get("ok") is True
        # Verify gone
        rl = requests.get(f"{API}/sauna/prices/import/history?limit=100",
                          headers=admin_headers, timeout=30)
        ids = [it["id"] for it in rl.json()["items"]]
        assert hid not in ids

    def test_delete_unknown_returns_404(self, admin_headers):
        d = requests.delete(f"{API}/sauna/prices/import/history/no-such-id-xyz",
                            headers=admin_headers, timeout=30)
        assert d.status_code == 404
